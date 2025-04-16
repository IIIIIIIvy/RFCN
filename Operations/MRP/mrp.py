import pandas as pd
from tkinter import messagebox
import tkinter as tk
import os
import datetime
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font,Alignment, Border,Side,PatternFill
from openpyxl.formatting.rule import CellIsRule
import warnings
from pandas.errors import SettingWithCopyWarning
warnings.simplefilter(action='ignore', category=(SettingWithCopyWarning))
pd.set_option('future.no_silent_downcasting', True)

file_name_dict={
    'sku_list':'SKU list.xlsx',
    'onhand_inventory':'On hand Inventory.xlsx',
    'fcst':'Forecast.xlsx',
    'op':'Confirmed OP.xlsx',
    'po':'Open Retail PO Qty.xlsx'
}

config_dict={
    'REG':{
        'On hand Inventory.xlsx':['REG'],
        'Forecast.xlsx':['REG'],
        'Confirmed OP.xlsx':['REGAL WAREHOUSE'],
        'Open Retail PO Qty.xlsx':['REG']
    },
    'RF':{
        'On hand Inventory.xlsx':['RF','MELR'],
        'Forecast.xlsx':['RF'],
        'Confirmed OP.xlsx':['RADIO FLYER'],
        'Open Retail PO Qty.xlsx':['RF']
    }
}

category_dict={
    0:'Confirmed OP',
    1:'Planned OP (due date)',
    2:'Open Retail PO Qty',
    3:'Bal. Fcst Qty',
    4:'Month end inventory'+'\n'+'(Deduct PO,FCST, SS)',
    5:'Upload JDE Forecast'+'\n'+'(Confirmed OP+Planned OP)'
}

year_now=datetime.now().year
month_now=datetime.now().month


def read_sku_list(folder_dir,file_name,target_name):
    sku_df=pd.read_excel(folder_dir+file_name,dtype=str,sheet_name=target_name)

    sku_df.fillna({'Safety Stock':0},inplace=True)
    sku_df.fillna({'Item Status':''},inplace=True)
    sku_df.fillna({'Est. Production Leadtime':0},inplace=True)
    sku_df.fillna({'Last Year Actuals Qty':0},inplace=True)
    sku_df.fillna({'Pcs/ctn':0},inplace=True)
    sku_df.fillna({'Remove flag':''},inplace=True)
    sku_df['Safety Stock']=sku_df['Safety Stock'].astype(int)
    sku_df['Est. Production Leadtime']=sku_df['Est. Production Leadtime'].astype(int)
    sku_df['Last Year Actuals Qty']=sku_df['Last Year Actuals Qty'].astype(int)
    sku_df['Pcs/ctn']=sku_df['Pcs/ctn'].astype(int)

    sku_df=sku_df[sku_df['Remove flag']!='Y']
    
    # ---- list needed    
    sku_list=sku_df['Item Number'].to_list()
    no_calculate_sku_list=sku_df[sku_df['Remove flag']=='P']['Item Number'].to_list()
    ss_dict=sku_df[['Item Number','Safety Stock']].set_index('Item Number').to_dict()['Safety Stock']
    factory_dict=sku_df[['Item Number','Factory']].set_index('Item Number').to_dict()['Factory']
    status_dict=sku_df[['Item Number','Item Status']].set_index('Item Number').to_dict()['Item Status']
    EPL_dict=sku_df[['Item Number','Est. Production Leadtime']].set_index('Item Number').to_dict()['Est. Production Leadtime']
    last_year_actuals_QTY_dict=sku_df[['Item Number','Last Year Actuals Qty']].set_index('Item Number').to_dict()['Last Year Actuals Qty']
    pcs_dict=sku_df[['Item Number','Pcs/ctn']].set_index('Item Number').to_dict()['Pcs/ctn']


    print("Successfully read data from SKU_List.")
    return sku_list,no_calculate_sku_list,ss_dict,factory_dict,EPL_dict,status_dict,last_year_actuals_QTY_dict,pcs_dict


def read_onhand_inventory(folder_dir,file_name,target_name,sku_list):
    on_hand_inventory_data=pd.read_excel(folder_dir+file_name,dtype=str)
    on_hand_inventory_data['Branch']=on_hand_inventory_data['Branch'].apply(lambda x:x.replace(' ','').upper())
    on_hand_inventory_data=on_hand_inventory_data[on_hand_inventory_data['Branch'].isin(config_dict[target_name][file_name])]
    on_hand_inventory_data=on_hand_inventory_data[on_hand_inventory_data['Item Number'].isin(sku_list)]
    on_hand_inventory_data.reset_index(drop=True,inplace=True)

    on_hand_inventory_data['Qty Onhand']=on_hand_inventory_data['Qty Onhand'].astype(int)

    inventory_res=on_hand_inventory_data.groupby('Item Number')['Qty Onhand'].sum().reset_index(name='onhand_qty_sum')
    # inventory_res=pd.merge(inventory_res,sku_df[['Item Number']],how='outer')
    inventory_res.fillna({'onhand_qty_sum':0},inplace=True)
    
    inventory_dict=inventory_res.set_index('Item Number').to_dict()['onhand_qty_sum']
    
    print("Successfully read data from On_Hand_Inventory.")
    return inventory_dict


def read_fcst(folder_dir,file_name,target_name,sku_list):
    # --- 初步计算fcst
    fcst_1=pd.read_excel(folder_dir+file_name,dtype=str,sheet_name=0,header=0)
    fcst_2=pd.read_excel(folder_dir+file_name,dtype=str,sheet_name=1,header=0)
    forecast=pd.concat([fcst_1,fcst_2[1:]],axis=0)
    forecast.fillna('',inplace=True)

    index=[x for x in range(12)]
    month=list(filter(lambda x:x.find('Unnamed')==-1,forecast.columns.tolist()))[1:]
    ori_month_dict=dict(zip(index,month))

    forecast.columns=forecast.iloc[0]
    forecast.drop(0,inplace=True)
    new_index=list(forecast.columns)
    for i in range(0,len(forecast.columns[6:])):
        forecast[forecast.columns[6+i]]=forecast[forecast.columns[6+i]].astype(int)
        if int((i)/4)>=12:
            new_index[6+i]=new_index[6+i]+'_'+str(year_now+1)+' '+ori_month_dict.get(int((i)/4)%12)
        else:
            new_index[6+i]=new_index[6+i]+'_'+str(year_now)+' '+ori_month_dict.get(int((i)/4))

    forecast=forecast.set_axis(new_index,axis='columns')

    forecast=forecast[forecast['Branch Plant'].apply(lambda x:x.replace(" ",'')).isin(config_dict[target_name][file_name])]
    forecast=forecast[forecast['Model'].isin(sku_list)]
    forecast.reset_index(drop=True,inplace=True)
    
    variance_temp_df=pd.DataFrame()
    for i in range(6,len(forecast.columns),4):
        col=forecast.columns[i].split('_')[1]
        variance_temp_df[col]=forecast[forecast.columns[i+2]]-forecast[forecast.columns[i+1]]-forecast[forecast.columns[i]]
        variance_temp_df[col]=variance_temp_df[col].apply(lambda x:x if x>0 else 0)
    variance_temp_df.columns=['Variance_'+col for col in list(variance_temp_df.columns)]

    for col in variance_temp_df.columns:
        forecast[col]=variance_temp_df[col]
    
    # --- datetime_col确定
    new_month_dict=ori_month_dict.copy()
    for key in new_month_dict.keys().__reversed__():
        new_key=key+1
        new_month_dict[new_key]=new_month_dict.pop(key)

    new_month_dict=dict(sorted(new_month_dict.items(),key=lambda x:x[0]))
    datetime_col=list(new_month_dict.values())
    datetime_col=[str(year_now)+' '+x for x in datetime_col[month_now-1:]]+[str(year_now+1)+' '+x for x in datetime_col[:month_now-1]]
    
    # --- 修正Fcst
    forecast_res=forecast.groupby(['Model'])[list(variance_temp_df.columns)].sum().reset_index()
    forecast_res=pd.concat([forecast_res['Model'],forecast_res.iloc[:,month_now:month_now+12]],axis=1)

    if forecast_res.shape[1]<13:
        col_len=forecast_res.shape[1]-1
        forecast_res=pd.concat([forecast_res,pd.DataFrame(columns=datetime_col[col_len:])],axis=1)

    forecast_res.rename(columns={'Model':'item_no'},inplace=True)
    forecast_res=pd.merge(forecast_res,pd.DataFrame(columns=['Item Number'],data=sku_list),left_on='item_no',right_on='Item Number',how='outer')
    forecast_res['item_no']=forecast_res['Item Number']
    del forecast_res['Item Number']
    forecast_res.fillna(0,inplace=True)
    
    print("Successfully read data from Forecast.")
    return forecast_res,new_month_dict,datetime_col


def read_confirmed_op(folder_dir,file_name,target_name,sku_list,new_month_dict,datetime_col):
    op_data_df=pd.read_excel(folder_dir+file_name,dtype=str)
    op_data_df=op_data_df[op_data_df['Ship To Name'].isin(config_dict[target_name][file_name])]
    op_data_df=op_data_df[op_data_df['2nd Item Number'].isin(sku_list)]

    op_data_df['Open Qty']=op_data_df['Open Qty'].astype('int')
    op_data_df['US Due Date']=op_data_df['US Due Date'].astype('datetime64[ns]')
    op_data_df.reset_index(drop=True,inplace=True)

    #--- 计算未来和overdue
    overdue_op_df=op_data_df[op_data_df['US Due Date'].apply(lambda x:x<datetime(year_now,month_now,1))]
    overdue_op_df.reset_index(drop=True,inplace=True)

    now_and_future_op_df=op_data_df[op_data_df['US Due Date'].apply(lambda x:x>=datetime(year_now,month_now,1))]
    now_and_future_op_df.reset_index(drop=True,inplace=True)

    now_and_future_op_df['Year of US Due Date']=now_and_future_op_df['US Due Date'].apply(lambda x:x.year)
    now_and_future_op_df['Month of US Due Date']=now_and_future_op_df['US Due Date'].apply(lambda x:x.month)

    now_and_future_op_sum_df=now_and_future_op_df.groupby(['2nd Item Number','Year of US Due Date','Month of US Due Date'])['Open Qty'].sum().reset_index(name='open_qty_sum')
    now_and_future_op_sum_df.sort_values(['2nd Item Number','Year of US Due Date','Month of US Due Date'],inplace=True)
    now_and_future_op_sum_df['Month of US Due Date']=now_and_future_op_sum_df['Month of US Due Date'].apply(lambda x:new_month_dict[x])
    now_and_future_op_sum_df['time']=now_and_future_op_sum_df['Year of US Due Date'].astype(str)+' '+now_and_future_op_sum_df['Month of US Due Date']
    del now_and_future_op_sum_df['Year of US Due Date'],now_and_future_op_sum_df['Month of US Due Date']
    
    overdue_op_sum_df=overdue_op_df.groupby(['2nd Item Number'])['Open Qty'].sum().reset_index(name='overdue_open_qty_sum')
    overdue_op_sum_df.sort_values(['2nd Item Number'],inplace=True)
    
    op_res=pd.merge(now_and_future_op_sum_df,overdue_op_sum_df,how='outer')
    op_res.fillna(0,inplace=True)
    
    # --- final result
    op_res_final=pd.DataFrame(columns=['item_no','overdue_op_qty']+datetime_col)
    pointer=0
    for index,rows in op_res.iterrows():
        op_res_final.loc[pointer,'item_no']=rows['2nd Item Number']
        op_res_final.loc[pointer,'overdue_op_qty']=rows['overdue_open_qty_sum']
        pointer+=1

        item_no=rows['2nd Item Number']
        col=rows['time']
        qty=rows['open_qty_sum']
        if col!=0:
            op_res_final.loc[op_res_final['item_no']==item_no,col]=qty

    op_res_final.drop_duplicates(['item_no'],inplace=True)
    op_res_final.reset_index(drop=True,inplace=True)
    
    op_res_final=pd.merge(op_res_final,pd.DataFrame(columns=['Item Number'],data=sku_list),left_on='item_no',right_on='Item Number',how='outer')
    op_res_final['item_no']=op_res_final['Item Number']
    del op_res_final['Item Number']
    op_res_final.fillna(0,inplace=True)
    
    print("Successfully read data from Confirmed_OP.")
    return op_res_final


def read_open_retail_po(folder_dir,file_name,target_name,sku_list,new_month_dict,datetime_col):
    po_data_df=pd.read_excel(folder_dir+file_name,dtype=str)
    po_data_df['Business Unit']=po_data_df['Business Unit'].apply(lambda x:x.upper())
    po_data_df=po_data_df[po_data_df['Business Unit'].isin(config_dict[target_name][file_name])]

    po_data_df['Next Status']=po_data_df['Next Status'].astype('int')
    po_data_df=po_data_df[(po_data_df['Next Status']<565)]
    po_data_df=po_data_df[po_data_df['2nd Item Number'].isin(sku_list)]

    po_data_df['Quantity']=po_data_df['Quantity'].astype('int')
    po_data_df['First Ship Date']=po_data_df['First Ship Date'].astype('datetime64[ns]')
    po_data_df.reset_index(drop=True,inplace=True)

    #--- 计算未来和overdue
    overdue_po_df=po_data_df[po_data_df['First Ship Date'].apply(lambda x:x<datetime(year_now,month_now,1))]
    overdue_po_df.reset_index(drop=True,inplace=True)

    now_and_future_po_df=po_data_df[po_data_df['First Ship Date'].apply(lambda x:x>=datetime(year_now,month_now,1))]
    now_and_future_po_df.reset_index(drop=True,inplace=True)    
    
    now_and_future_po_df['Year of First Ship Date']=now_and_future_po_df['First Ship Date'].apply(lambda x:x.year)
    now_and_future_po_df['Month of First Ship Date']=now_and_future_po_df['First Ship Date'].apply(lambda x:x.month)

    now_and_future_po_sum_df=now_and_future_po_df.groupby(['2nd Item Number','Year of First Ship Date','Month of First Ship Date'])['Quantity'].sum().reset_index(name='qty_sum')
    now_and_future_po_sum_df.sort_values(['2nd Item Number','Year of First Ship Date','Month of First Ship Date'],inplace=True)

    now_and_future_po_sum_df['Month of First Ship Date']=now_and_future_po_sum_df['Month of First Ship Date'].apply(lambda x:new_month_dict[x])
    now_and_future_po_sum_df['time']=now_and_future_po_sum_df['Year of First Ship Date'].astype(str)+' '+now_and_future_po_sum_df['Month of First Ship Date']
    del now_and_future_po_sum_df['Year of First Ship Date'],now_and_future_po_sum_df['Month of First Ship Date']
    
    overdue_po_sum_df=overdue_po_df.groupby(['2nd Item Number'])['Quantity'].sum().reset_index(name='overdue_qty_sum')
    overdue_po_sum_df.sort_values(['2nd Item Number'],inplace=True)
    
    po_res=pd.merge(now_and_future_po_sum_df,overdue_po_sum_df,how='outer')
    po_res.fillna(0,inplace=True)
    
    # --- final result
    po_res_final=pd.DataFrame(columns=['item_no','overdue_po_qty']+datetime_col)
    pointer=0
    for index,rows in po_res.iterrows():
        po_res_final.loc[pointer,'item_no']=rows['2nd Item Number']
        po_res_final.loc[pointer,'overdue_po_qty']=rows['overdue_qty_sum']
        pointer+=1

        item_no=rows['2nd Item Number']
        col=rows['time']
        qty=rows['qty_sum']

        if col!=0:
            po_res_final.loc[po_res_final['item_no']==item_no,col]=qty

    po_res_final.drop_duplicates(['item_no'],inplace=True)
    po_res_final.reset_index(drop=True,inplace=True)
    
    po_res_final=pd.merge(po_res_final,pd.DataFrame(columns=['Item Number'],data=sku_list),left_on='item_no',right_on='Item Number',how='outer')
    po_res_final['item_no']=po_res_final['Item Number']
    del po_res_final['Item Number']
    po_res_final.fillna(0,inplace=True)
    
    print("Successfully read data from Open_Retail_PO_Qty.")
    return po_res_final


def get_write_df(target_name,datetime_col,dict_from_sku_list,inventory_dict,forecast_res,op_res_final,po_res_final):
    sku_list=dict_from_sku_list[0]
    no_calculate_sku_list=dict_from_sku_list[1]
    ss_dict=dict_from_sku_list[2]
    factory_dict=dict_from_sku_list[3]
    EPL_dict=dict_from_sku_list[4]
    status_dict=dict_from_sku_list[5]
    last_year_actuals_QTY_dict=dict_from_sku_list[6]
    pcs_dict=dict_from_sku_list[7]

    template_df=pd.DataFrame(columns=['SKU','Factory','On hand Inventory','Safey stock','Category','Overdue month']+
                         datetime_col+['Pcs/ctn','Item Status','Est. Production Leadtime', str(datetime.now().year-1)+' Actuals Qty'])
    for sku in sku_list:
        row_index=len(template_df)
        for i in range(0,6):
            template_df.loc[row_index+i,'SKU']=sku
            template_df.loc[row_index+i,'On hand Inventory']=inventory_dict[sku]
            template_df.loc[row_index+i,'Safey stock']=ss_dict[sku]
            template_df.loc[row_index+i,'Factory']=factory_dict[sku]
            template_df.loc[row_index+i,'Pcs/ctn']=pcs_dict[sku]
            template_df.loc[row_index+i,'Item Status']=status_dict[sku]
            template_df.loc[row_index+i,'Est. Production Leadtime']=EPL_dict[sku]
            template_df.loc[row_index+i,str(datetime.now().year-1)+' Actuals Qty']=last_year_actuals_QTY_dict[sku]

            template_df.loc[row_index+i,'Category']=category_dict[i]

    # --- op
    op_start_col=5
    for i in range(0,len(template_df),6):
        sku=template_df.loc[i,'SKU']
        op_temp=op_res_final.loc[op_res_final['item_no']==sku]
        
        template_df.iloc[i,op_start_col:op_start_col+13]=op_temp.iloc[0,1:14]

    # --- po
    po_start_col=5
    for i in range(2,len(template_df),6):
        sku=template_df.loc[i,'SKU']
        po_temp=po_res_final.loc[po_res_final['item_no']==sku]
        
        template_df.iloc[i,po_start_col:po_start_col+13]=po_temp.iloc[0,1:14]

    # --- fcst
    fcst_start_col=6
    for i in range(3,len(template_df),6):
        sku=template_df.loc[i,'SKU']
        fcst_temp=forecast_res.loc[forecast_res['item_no']==sku]
        
        template_df.iloc[i,fcst_start_col:fcst_start_col+12]=fcst_temp.iloc[0,1:14]

    fill_flag_df=template_df[(template_df['Category']=='Open Retail PO Qty') | (template_df['Category']=='Bal. Fcst Qty')][['SKU']+datetime_col]
    no_fill_sku_list=[]
    for sku in sku_list:
        fill_flag_temp_df=fill_flag_df[fill_flag_df['SKU']==sku]
        all_zero_flag=True
        for col in datetime_col:
            fill_flag_temp_df_2=fill_flag_temp_df[datetime_col]==0
            all_zero_flag=fill_flag_temp_df_2[col].all()
            if ~all_zero_flag:
                break
        if all_zero_flag:
            no_fill_sku_list.append(sku)

    # ----history planned op
    file_path_list = Path(folder_dir+'\\HISTORY DOCUMENTS\\').glob('Supply plan_'+target_name+'_*[0-9].xlsx') 
    file_name_list=[x.name for x in file_path_list]
    file_name_list.sort()
    latest_file_name=file_name_list[-1]
    history_df=pd.read_excel(folder_dir+'\\HISTORY DOCUMENTS\\'+latest_file_name)
    history_df['SKU']=history_df['SKU'].astype(str)
    # history_df=pd.DataFrame(columns=['SKU','Factory','On hand Inventory','Safey stock','Category','Overdue month']+datetime_col+['Item Status',
    #        'Est. Production Leadtime', '2023 Actuals Qty','Notes'])

    history_planned_po_df=history_df[history_df['Category']=='Planned OP (due date)']
    history_datetime_col=list(history_planned_po_df.columns[6:18])
    history_planned_po_df=history_planned_po_df.loc[:,['SKU','Category']+history_datetime_col]
    history_planned_po_df.fillna(0,inplace=True)

    history_planned_po_df=history_planned_po_df.merge(template_df,on=['SKU','Category'],how='right')
    overlap_col=[x  for x in history_planned_po_df.columns if x[-2:]=='_x']
    history_planned_po_df=history_planned_po_df[['SKU']+overlap_col]

    pl_op_start_col=6
    for i in range(1,len(template_df),6):
        sku=template_df.loc[i,'SKU']
        
        pl_op_temp=history_planned_po_df.loc[history_planned_po_df['SKU']==sku]
        if len(pl_op_temp)!=0:
            template_df.iloc[i,pl_op_start_col:pl_op_start_col+len(overlap_col)]=pl_op_temp.iloc[1,1:]

    template_df.loc[template_df['Category']=='Planned OP (due date)',datetime_col]=template_df.loc[template_df['Category']=='Planned OP (due date)',datetime_col].fillna(0)

    # --- notes
    history_notes_df=history_df[['SKU','Notes']]

    history_notes_df.drop_duplicates(['SKU'],keep='first',inplace=True)
    history_notes_df.rename(columns={'Notes':'History Notes'},inplace=True)

    template_df=template_df.merge(history_notes_df,on='SKU',how='left')
    template_df['Notes']=template_df['History Notes']
    del template_df['History Notes']

    # --- first month end ivt
    calculation_start_col=6
    for i in range(4,len(template_df),6):
        ivt=template_df.loc[i,'On hand Inventory']
        ss=template_df.loc[i,'Safey stock']
        ovd_op=template_df.loc[i-4,'Overdue month']
        ovd_po=template_df.loc[i-2,'Overdue month']

        for col in range(calculation_start_col,calculation_start_col+12):
            curr_op=template_df.iloc[i-4,col]
            curr_po=template_df.iloc[i-2,col]
            curr_fcst=template_df.iloc[i-1,col]
            curr_pl_op=template_df.iloc[i-3,col]
            if curr_pl_op==0:
                template_df.iloc[i-3,col]=''
                
            if col==calculation_start_col:
                curr_me_ivt=ivt - ss + ovd_op - ovd_po + curr_op + curr_pl_op - curr_fcst - curr_po
            else:
                past_ivt = template_df.iloc[i,col-1]
                curr_me_ivt = past_ivt + curr_op + curr_pl_op - curr_fcst - curr_po

            if col in [6,7,8]:
                if curr_me_ivt<0:
                    curr_me_ivt=0
            
            template_df.iloc[i,col]=curr_me_ivt

    # --- 计算planned op并修正month end ivt
    calculate_sku_list=[]
    calculation_start_col=9 # j列开始计算
    for i in range(1,len(template_df),6):
        if template_df.loc[i,'SKU'] in no_calculate_sku_list or  template_df.loc[i,'SKU'] in no_fill_sku_list:
            continue
        
        ivt=template_df.loc[i,'On hand Inventory']
        ss=template_df.loc[i,'Safey stock']
        ovd_op=template_df.loc[i-1,'Overdue month']
        ovd_po=template_df.loc[i+1,'Overdue month']

        # 遍历找到ivt<50的列index
        end_col=calculation_start_col+2
        while end_col<18:
            start_col=end_col
            while start_col<18:
                target_ivt=template_df.iloc[i+3,start_col]
                if target_ivt<50:
                    calculate_sku_list.append(template_df.loc[i,'SKU'])
                    break
                else:
                    start_col+=1

            # 计算planned op
            # for col in range(start_col-2,17):
            if start_col<18:
                col= start_col - 2
                curr_op=template_df.iloc[i-1,col]
                curr_po=template_df.iloc[i+1,col]
                curr_fcst=template_df.iloc[i+2,col]
                
                next_op_1=template_df.iloc[i-1,col+1]
                next_po_1=template_df.iloc[i+1,col+1]
                next_fcst_1=template_df.iloc[i+2,col+1]
                next_pl_op_1=0

                next_op_2=template_df.iloc[i-1,col+2]
                next_po_2=template_df.iloc[i+1,col+2]
                next_fcst_2=template_df.iloc[i+2,col+2]
                next_pl_op_2=0

                past_ivt=template_df.iloc[i+3,col-1]

                curr_pl_op=0
                while target_ivt<50 :
                    curr_pl_op+=50 # 步长为50
                    # 重新计算各ivt

                    # 第一列ivt
                    if col==6:
                        curr_ivt=ivt - ss + ovd_op - ovd_po + curr_op + curr_pl_op - curr_fcst - curr_po
                    else:
                        curr_ivt=past_ivt + curr_op + curr_pl_op - curr_fcst - curr_po

                    # 第二列ivt
                    next_ivt_1 = curr_ivt + next_op_1 + next_pl_op_1 - next_fcst_1 - next_po_1
                    # 第三列ivt
                    next_ivt_2 = next_ivt_1 + next_op_2 + next_pl_op_2 - next_fcst_2 - next_po_2
                    target_ivt = next_ivt_2

                # 写回计算数据：planned op
                template_df.iloc[i,col]=curr_pl_op
                template_df.iloc[i+3,col]=curr_ivt
                template_df.iloc[i+3,col+1]=next_ivt_1
                template_df.iloc[i+3,col+2]=next_ivt_2
                for j in range(col+3,18):
                    curr_op=template_df.iloc[i-1,j]
                    curr_po=template_df.iloc[i+1,j]
                    curr_fcst=template_df.iloc[i+2,j]
                    curr_pl_op=0

                    past_ivt=template_df.iloc[i+3,j-1]
                    template_df.iloc[i+3,j]=past_ivt + curr_op + curr_pl_op - curr_fcst - curr_po
                end_col=col+5
            else:
                break
            

            
    calculate_sku_list=list(set(calculate_sku_list))
    calculate_sku_list.sort()

    return template_df,calculate_sku_list,no_fill_sku_list


def write_new_file(folder_dir,target_name,template_df,calculate_sku_list,no_fill_sku_list,no_calculate_sku_list):
    document_name='Supply plan_'+target_name+'_'+str(month_now).zfill(2)+str(datetime.now().day).zfill(2)+'.xlsx'
    writer = pd.ExcelWriter(folder_dir+document_name, engine='openpyxl')

    template_df.to_excel(writer, index=False)
    worksheet = writer.sheets['Sheet1']

    worksheet.column_dimensions['A'].width=18.75
    worksheet.column_dimensions['B'].width=13.75
    worksheet.column_dimensions['C'].width=7.75
    worksheet.column_dimensions['D'].width=7.75
    worksheet.column_dimensions['E'].width=30.75
    worksheet.column_dimensions['F'].width=11.5
    worksheet.column_dimensions['S'].width=5.7
    worksheet.column_dimensions['T'].width=13.7
    worksheet.column_dimensions['U'].width=14
    worksheet.column_dimensions['V'].width=11.5
    worksheet.column_dimensions['W'].width=40.7
    for col in range(ord('G'),ord('R')+1):
        worksheet.column_dimensions[chr(col)].width=7.75

    # 冻结窗格
    worksheet.freeze_panes='G2'

    # -------------------- format ---------------------------------------
    # 表头
    for cell in worksheet[1]:
        cell.font=Font(name="Calibri", size=13,bold=True)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=Border(bottom=Side(style='medium'))
        if cell.column_letter =='A':
            cell.fill=PatternFill(start_color='DAEEF3',end_color='DAEEF3',fill_type='solid')
        if cell.column_letter == 'D'  or cell.column_letter == 'R':
            cell.border=Border(right=Side(style='medium'))
            

    # 所有行（表格内容）字体设置
    for i in range(2, len(template_df)+2):
        worksheet.row_dimensions[i].height=32
        for cell in worksheet[i]:
            if cell.column_letter == 'A':
                cell.font=Font(name="Calibri", size=13,bold=True)
                cell.alignment=Alignment(horizontal='center',vertical='center')
                cell.fill=PatternFill(start_color='DAEEF3',end_color='DAEEF3',fill_type='solid')
            elif cell.column_letter == 'E':
                if (i-3)%6==0 or (i-6)%6==0:
                    cell.font=Font(name="Calibri", size=11,color='0000FF',bold=True)
                else:
                    cell.font=Font(name="Calibri", size=11,bold=True)

                cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            elif cell.column_letter == 'T':
                cell.font=Font(name="Calibri", size=12)
            else:
                cell.font=Font(name="Calibri", size=13)
                cell.alignment=Alignment(horizontal='center',vertical='center')

            if cell.column_letter == 'D'  or cell.column_letter == 'R':
                cell.border=Border(right=Side(style='medium'))

            if cell.column_letter == 'C' and cell.value==0:
                cell.fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
        
    for i in range(7, len(template_df)+2,6):
        for cell in worksheet[i]:
            if cell.column_letter == 'D'  or cell.column_letter == 'R':
                cell.border=Border(bottom=Side(style='medium'),right=Side(style='medium'))
            else:
                cell.border=Border(bottom=Side(style='medium'))
    
    for i in range(2,len(template_df)+2,6):
        for col in ['C','D','S','T','U','V','W']:
            worksheet.merge_cells(col+str(i)+':'+col+str(i+5))
            if col=='S' and worksheet[col+str(i)].value>1:
                worksheet[col+str(i)].fill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
            elif col=='W' or col=='T':
                worksheet[col+str(i)].alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            # worksheet[col+str(i)].font=Font(name="Calibri", size=13,bold=True)

        
    # ------------------- formula ---------------------------------------
    # 条件格式
    redFill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    pinkFill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid')
    yellowFill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    redFont=Font(name="Calibri", size=13,color='9C0006')
    rule1=CellIsRule(operator='lessThan', formula=['50'], stopIfTrue=False, fill=pinkFill)
    rule3=CellIsRule(operator='equal', formula=['0'], stopIfTrue=False, fill=yellowFill)

            
    for writer_pointer in range(6,len(template_df)+2,6):
        # ----month end inventory
        ori_formula='C'+str(writer_pointer-4)+'+G'+str(writer_pointer-4)+'+F'+str(writer_pointer-4)+'+G'+str(writer_pointer-3)+'-F'+str(writer_pointer-2)+'-G'+str(writer_pointer-2)+'-G'+str(writer_pointer-1)+'-D'+str(writer_pointer-4)
        worksheet['G'+str(writer_pointer)]='=IF('+ori_formula+'<0,0,'+ori_formula+')'

        col_char='H'
        while col_char<'S':
            pre_col_char=chr(ord(col_char)-1)
            ori_formula=pre_col_char+str(writer_pointer)+'+'+col_char+str(writer_pointer-4)+'+'+col_char+str(writer_pointer-3)+'-'+col_char+str(writer_pointer-2)+'-'+col_char+str(writer_pointer-1)
            if col_char=='H' or col_char=='I':
                # worksheet[col_char+str(writer_pointer)]='=IF(AND(G'+str(writer_pointer)+'>0,'+ori_formula+'<0),0,'+ori_formula+')'
                worksheet[col_char+str(writer_pointer)]='=IF('+ori_formula+'<0,0,'+ori_formula+')'
            # elif col_char=='I':
            #     worksheet[col_char+str(writer_pointer)]='=IF(AND(G'+str(writer_pointer)+'>0,H'+str(writer_pointer)+'>0,'+ori_formula+'<0),0,'+ori_formula+')'
            else:
                worksheet[col_char+str(writer_pointer)]='='+ori_formula
            
            col_char=chr(ord(col_char)+1)

        col_char='G'
        while col_char<'R':
            next_col_char=chr(ord(col_char)+1)
            rule2=CellIsRule(operator='lessThan', formula=['$'+next_col_char+'$'+str(writer_pointer-1)+'+$'+next_col_char+'$'+str(writer_pointer-2)], stopIfTrue=False, fill=redFill)
            worksheet.conditional_formatting.add('$'+col_char+'$'+str(writer_pointer),rule2)
            col_char=chr(ord(col_char)+1)  

        worksheet.conditional_formatting.add('$R$'+str(writer_pointer),rule1)

        # ----REGAL JDE Fcst
        col_char='F'
        while col_char<'R':
            next_col_char=chr(ord(col_char)+1)
            worksheet[col_char+str(writer_pointer+1)]='='+next_col_char+str(writer_pointer-4)+'+'+next_col_char+str(writer_pointer-3)
            col_char=next_col_char
        worksheet['R'+str(writer_pointer+1)].value=0
            

        # ----planned op
        sku=worksheet['A'+str(writer_pointer)].value
        if sku in calculate_sku_list:
            col_char='G'
            while col_char<'S':
                worksheet[col_char+str(writer_pointer-3)].fill=PatternFill(start_color='FDE9D9',end_color='FDE9D9',fill_type='solid')
                col_char=chr(ord(col_char)+1)

        if sku in no_fill_sku_list:
            col_char='G'
            while col_char<'S':
                worksheet[col_char+str(writer_pointer-3)].fill=PatternFill(start_color='BFBFBF',end_color='BFBFBF',fill_type='solid')
                # worksheet[col_char+str(writer_pointer-3)]=''
                col_char=chr(ord(col_char)+1)

        if sku in no_calculate_sku_list:
            col_char='G'
            while col_char<'S':
                worksheet[col_char+str(writer_pointer-3)].fill=PatternFill(start_color='000000',end_color='000000',fill_type='solid')
                worksheet[col_char+str(writer_pointer-3)]=''
                col_char=chr(ord(col_char)+1)


    # ------------------------------------------Planning schedule------------------------------------------------
    schedule_sheet_name='Planning schedule'
    pd.DataFrame(columns=['Planned date','CRD','US due date']).to_excel(writer, index=False,sheet_name=schedule_sheet_name)
    worksheet = writer.sheets[schedule_sheet_name]
    worksheet.column_dimensions['A'].width=16
    worksheet.column_dimensions['B'].width=13
    worksheet.column_dimensions['C'].width=15

    us_due_date_dict={'RF':44,'REG':30}
    worksheet['A2']='=NOW()'
    worksheet['A3']='=A2+7'
    worksheet['A4']='=A3+7'

    worksheet['B2']='=IF(AND(MONTH(A2+45)<10,MONTH(A2+45)>5),A2+60,A2+45)'
    worksheet['B3']='=IF(AND(MONTH(A3+45)<10,MONTH(A3+45)>5),A3+60,A3+45)'
    worksheet['B4']='=IF(AND(MONTH(A4+45)<10,MONTH(A4+45)>5),A4+60,A4+45)'

    worksheet['C2']='=B2+'+str(us_due_date_dict[target_name])
    worksheet['C3']='=B3+'+str(us_due_date_dict[target_name])
    worksheet['C4']='=B4+'+str(us_due_date_dict[target_name])

    # 所有行（表格内容）字体设置
    for i in range(1, 5):
        worksheet.row_dimensions[i].height=20
        for cell in worksheet[i]:
            if i == 1:
                cell.font=Font(name="Calibri", size=13,bold=True)
            else:
                cell.font=Font(name="Calibri", size=13)
                cell.number_format='m/d/yyyy'
            cell.alignment=Alignment(horizontal='center',vertical='center')

    writer.close()


if __name__ == '__main__':
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    messagebox.showinfo("提示", "程序启动中！")

    # folder_dir = os.getcwd()+"\\MRP"+"\\"
    folder_dir = os.getcwd()+"\\"

    for target_name in ['RF','REG']:
        print('--------------------------------------')
        sku_list, no_calculate_sku_list, ss_dict, factory_dict, EPL_dict, status_dict, last_year_actuals_QTY_dict, pcs_dict = read_sku_list(folder_dir,file_name_dict['sku_list'],target_name)
        inventory_dict = read_onhand_inventory(folder_dir,file_name_dict['onhand_inventory'],target_name,sku_list)
        forecast_res, new_month_dict, datetime_col = read_fcst(folder_dir,file_name_dict['fcst'],target_name,sku_list)
        op_res_final = read_confirmed_op(folder_dir,file_name_dict['op'],target_name,sku_list,new_month_dict,datetime_col)
        po_res_final = read_open_retail_po(folder_dir,file_name_dict['po'],target_name,sku_list,new_month_dict,datetime_col)

        dict_from_sku_list=[sku_list, no_calculate_sku_list, ss_dict, factory_dict, EPL_dict, status_dict, last_year_actuals_QTY_dict, pcs_dict]
        template_df, calculate_sku_list, no_fill_sku_list = get_write_df(target_name,datetime_col,dict_from_sku_list,inventory_dict,forecast_res,op_res_final,po_res_final)
        write_new_file(folder_dir,target_name,template_df,calculate_sku_list,no_fill_sku_list,no_calculate_sku_list)
        print('Finish writing '+target_name+'.')

    messagebox.showinfo("提示", "Completed.")