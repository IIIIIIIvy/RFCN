import pandas as pd
from tkinter import messagebox
import tkinter as tk
import os
import math
from openpyxl.styles import Font, Alignment, colors, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl import load_workbook
from datetime import datetime, timedelta
import warnings
from pandas.errors import SettingWithCopyWarning

warnings.simplefilter(action="ignore", category=(SettingWithCopyWarning))
# 修改日志：
# - 1/22/2015 修改bug：当某单刚好成一个柜子，即TTL CBM大于min又小于max时的错误
#

USER_CONFIG={
    'Lyn':{
        'SCALE_FOR_SUMMARY':47,
        'SCALE_FOR_DETAIL':53,
    },
    'Rainbow':{
        'SCALE_FOR_SUMMARY':50,
        'SCALE_FOR_DETAIL':56,
    },
}

CLP_DICT={
    '<66':'CFS',
    '>=66':'40HQ'
}

month_dict={1: 'Jan',
 2: 'Feb',
 3: 'Mar',
 4: 'Apr',
 5: 'May',
 6: 'Jun',
 7: 'Jul',
 8: 'Aug',
 9: 'Sep',
 10: 'Oct',
 11: 'Nov',
 12: 'Dec'}

number_to_ordinal_dict={
    1:'1ST',
    2:'2ND',
    3:'3RD',
    4:'4TH',
    5:'5TH',
    6:'6TH',
    7:'7TH',
    8:'8TH',
    9:'9TH',
    10:'10TH',
    11:'11TH',
    12:'12TH',
    13:'13TH',
    14:'14TH',
    15:'15TH',
    16:'16TH',
    17:'17TH',
    18:'18TH',
    19:'19TH',
    20:'20TH',
}

split_item_color_list=[
    'FCD5B4',
    'CCC0DA',
    'D8E4BC',
    '00FFFF',
    'CCFF66',
    'FF99FF',
    '3399FF',
    'FEE062',
    '00CC00',
    'CCECFF',
]


column_width_dict={
    'A':9,
    'B':8,
    'C':8,
    'D':8,
    'E':11,
    'F':9,
    'G':8,
    'H':21,
    'I':10,
    'J':12,
    'K':14,
    'M':11,
    'N':11,
    'O':11,
    'P':6,
    'U':13,
    'V':13,
    'W':13,
    'Y':7.8,
    'Z':7.8,
    'AA':7.8,
}

row_height_dict={
    'sheet_title':48,
    'table_content':20.5,
    'grey_content':7
}

detail_write_columns=['BK#', 'Pick up', 'ETD', 'E-doc', 'Order Number',
       'Related Order Number', 'Vendor Name', 'Customer PO', '2nd Item Number',
       'TCIN#', 'DPCI#', 'Quantity', 'First Ship Date', 'Last Ship Date',
       'Cargo Ready Date', 'Qty/\nCarton', 'Net Weight (kg)',
       'Gross Weight (kg)', 'Cubic\nMeters (per carton)', 'TTL CTNS',
       'TTL NW (KG)', 'TTL GW (KG)', 'TTL CBM','CLP', 
        'Width (L) cm', 'Depth (W) cm','Height (H) cm',  'groupIndex']
summary_write_columns=['BK#', 'Pick up', 'ETD', 'E-doc', 'Order Number',
       'Related Order Number', 'Vendor Name', 'Customer PO', '2nd Item Number',
       'TCIN#', 'DPCI#', 'Quantity', 'First Ship Date', 'Last Ship Date',
       'Cargo Ready Date', 'Qty/\nCarton', 'Net Weight (kg)',
       'Gross Weight (kg)', 'Cubic\nMeters (per carton)', 'TTL CTNS',
       'TTL NW (KG)', 'TTL GW (KG)', 'TTL CBM', 'CLP', 'Unit Price',
       'TTL Amount', 'Payment Term', 'Width (L) cm', 'Depth (W) cm',
       'Height (H) cm',  'groupIndex']


def splitCLP(ttl_cbm,temp_res):
    if ttl_cbm>=66:
        box_qty=int(ttl_cbm/66)
        temp_res+=str(box_qty)+'*'+CLP_DICT['>=66']
        ttl_cbm=ttl_cbm%66
    else:
        temp_res+=CLP_DICT['<66']
        ttl_cbm-=66

    temp_res+=','
    if ttl_cbm>0:
        temp_res = splitCLP(ttl_cbm,temp_res)
    return temp_res 



def shipping_window_overlap_calculation(data):
    # -----获取target，遍历判断时间窗口是否重叠
    groupBy = data.groupby(['Vendor Name','Customer PO',])
    groupIndex=1

    for name,group in groupBy:
        # print('-------------------')
        # print(name)
        temp = groupBy.get_group(name)
        if len(temp)>1:
            # print('BEFORE')
            # print(temp[['First Ship Date','Last Ship Date']],)
            
            # 判断时间窗口是否重叠：Sa≥Eb | Sb≥Ea，则不重叠；反之重叠。
            target_first_ship_date=temp.iloc[0]['First Ship Date']
            target_last_ship_date=temp.iloc[0]['Last Ship Date']

            inner_group_index=1
            for index,rows in temp.iterrows():
                if target_first_ship_date<rows['Last Ship Date']+timedelta(days=-1) and target_last_ship_date>rows['First Ship Date']+timedelta(days=1):
                    temp.loc[index,'overlap']=inner_group_index
                else:
                    target_first_ship_date=rows['First Ship Date']
                    target_last_ship_date=rows['Last Ship Date']
                    inner_group_index+=1
                    temp.loc[index,'overlap']=inner_group_index
                    
            # print('AFTER')
            # print(temp[['First Ship Date','Last Ship Date','overlap']])

            overlap_list=temp['overlap'].drop_duplicates().tolist()
            for flag in overlap_list:
                data.loc[temp[temp['overlap']==flag].index,'groupIndex']=groupIndex
                groupIndex+=1
                # print(data.loc[temp[temp['overlap']==flag].index][['Vendor Name','Customer PO','First Ship Date','Last Ship Date','groupIndex']])
                # print()
        else:
            data.loc[temp.index,'groupIndex']=groupIndex
            groupIndex+=1
            # print(data.loc[temp.index][['Vendor Name','Customer PO','First Ship Date','Last Ship Date','groupIndex']])
            # print()

    return data


def data_extraction(file_dir, data_file_name):
    # ------------------------------------------------------------monthly_data------------------------------------------------------------
    monthly_data=pd.read_excel(file_dir+data_file_name,sheet_name='Monthly Data',dtype=str)
    monthly_data=monthly_data.sort_values(['Customer PO','Vendor Name','First Ship Date','Last Ship Date'])
    monthly_data.reset_index(drop=True,inplace=True)
    monthly_data['Last Ship Date']=monthly_data['Last Ship Date'].apply(lambda x:datetime.strptime(x,'%Y-%m-%d %H:%M:%S'))
    monthly_data['First Ship Date']=monthly_data['First Ship Date'].apply(lambda x:datetime.strptime(x,'%Y-%m-%d %H:%M:%S'))
    monthly_data['Cargo Ready Date']=monthly_data['Cargo Ready Date'].apply(lambda x:datetime.strptime(x,'%Y-%m-%d %H:%M:%S'))

    change_column=['Quantity','Qty/\nCarton', 'Net Weight (kg)',
        'Gross Weight (kg)', 'Cubic\nMeters (per carton)', 'TTL CTNS',
        'TTL NW (KG)', 'TTL GW (KG)', 'TTL CBM','Unit Price','Width (L) cm',
        'Depth (W) cm', 'Height (H) cm']
    for column in change_column:
        monthly_data[column]=monthly_data[column].astype(float)
    monthly_data['Battery']=monthly_data['Battery'].apply(lambda x:x.lower())
    monthly_data['TTL CBM'] = monthly_data['TTL CBM'].apply(lambda x:round(x,3))

    monthly_data['BK#']=''
    monthly_data['Pick up']=''
    monthly_data['ETD']=''
    monthly_data['E-doc']=''

    monthly_data=monthly_data[[ 'BK#','Pick up','ETD','E-doc','Order Number', 'Related Order Number', 'Vendor Name',
        'Customer PO', '2nd Item Number', 'TCIN#',
        'ASIN# or SKU#... ', 'Quantity', 'First Ship Date', 'Last Ship Date',
        'Cargo Ready Date', 'Qty/\nCarton',
        'Net Weight (kg)', 'Gross Weight (kg)', 'Cubic\nMeters (per carton)',
        'TTL CTNS', 'TTL NW (KG)', 'TTL GW (KG)', 'TTL CBM', 'Battery','Unit Price', 'TTL Amount', 'Payment Term','Width (L) cm',
        'Depth (W) cm', 'Height (H) cm',]]
    monthly_data.rename(columns={'ASIN# or SKU#... ':'DPCI#'},inplace=True)
    
    # ------------------------------------------------------------battery_item------------------------------------------------------------
    battery_item_list=monthly_data[monthly_data['Battery']=='yes']['2nd Item Number'].drop_duplicates().tolist()
    sorted(battery_item_list)
    
    # ------------------------------------------------------------item_loading_qty------------------------------------------------------------
    item_loading_qty_df=pd.read_excel(file_dir+'TGT item loading qty.xlsx')
    item_loading_qty_df['Model #']=item_loading_qty_df['Model #'].astype(str)
    item_loading_qty_df['Model #']=item_loading_qty_df['Model #'].apply(lambda x:x.strip())
    item_loading_qty_df.sort_values(['Factory','Model #'],inplace=True)
    item_loading_qty_df.reset_index(drop=True,inplace=True)

    monthly_data=pd.merge(monthly_data,item_loading_qty_df[['Factory', 'Model #','Min loading qty','CBM*Min loading qty','Container Quantity (Piece)','CBM*Max loading qty' ]],left_on=['Vendor Name','2nd Item Number'],right_on=['Factory', 'Model #'],how='left')
    del monthly_data['Factory'],monthly_data['Model #']

    monthly_data = shipping_window_overlap_calculation(monthly_data)
    monthly_data.sort_values(['First Ship Date','groupIndex'],inplace=True)
    
    # ------------------------------------------------------------sheet_name_list------------------------------------------------------------
    sheet_name_list=[]
    for vendor in monthly_data['Vendor Name'].unique():
        sheet_name_list.append(vendor)
        sheet_name_list.append(vendor+' 分柜')

    return monthly_data, battery_item_list, sheet_name_list


def write_excel(monthly_data, battery_item_list, sheet_name_list, document_dir):
    document_name=str(datetime.now().month).zfill(2)+str(datetime.now().day).zfill(2)+'_RESULT'+'.xlsx'
    writer = pd.ExcelWriter(document_dir+document_name, engine='openpyxl')

    # 先写入空表，定sheet顺序
    for sheet_name in sheet_name_list:
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)

    # ------------------------------------分柜------------------------------------------------------------
    for sheet_name in sheet_name_list:
        if sheet_name.find(' 分柜')==-1:
            continue
        print('---------------------------------------------')
        print(sheet_name)
        
        worksheet = writer.sheets[sheet_name]
        
        for key,value in column_width_dict.items():
            worksheet.column_dimensions[key].width=value
        
        worksheet.column_dimensions['Y'].hidden=1
        worksheet.column_dimensions['Z'].hidden=1
        worksheet.column_dimensions['AA'].hidden=1
        worksheet.column_dimensions['AB'].hidden=1
        
        #----------------------content----------------------
        writer_pointer=0
        start_list=[]
        end_list=[]
        item_start_list=[]
        item_end_list=[]
        formula_list=[]
        formula_add_dict={}
        grey_list=[]

        # ----
        df=monthly_data[monthly_data['Vendor Name']==sheet_name[:-3]]
        worksheet['F1']='TGT - '+month_dict[df['First Ship Date'].sort_values().tolist()[0].month]+' - '+sheet_name[:-3]
        writer_pointer+=1

        first_group_flag=False
        for index in df['groupIndex'].unique():
            temp_formula_list_end=0
            if list(df['groupIndex'].unique()).index(index)==0:
                first_group_flag=True
            writer_pointer+=1
            groupdf=df[df['groupIndex']==index]
            # del groupdf['groupIndex']

            # -------先写CLP已经计算好的部分
            groupdf['split_flag']=groupdf['TTL CBM']>=groupdf['CBM*Min loading qty']
            groupdf['CLP']=None

            while groupdf['CLP'].hasnans:
                groupdf.sort_values(['TTL CBM','CLP'],ascending=[False,True],inplace=True)
                groupdf.reset_index(drop=True,inplace=True)
                row_pointer=groupdf[groupdf['CLP'].isna()].index[0]
        
                
                if groupdf.loc[row_pointer,'split_flag']:
                    groupdf.loc[row_pointer,'CLP']='1*40HQ'
                    ttl_cbm=groupdf.loc[row_pointer,'TTL CBM']
                    max_cbm=groupdf.loc[row_pointer,'CBM*Max loading qty']

                    if ttl_cbm>max_cbm:
                        # print(temp_df.loc[row_pointer,'2nd Item Number'])
                        # 存一个备份给下一个df,进行split
                        new_df=pd.DataFrame(groupdf.loc[row_pointer]).T
                        new_df.reset_index(drop=True,inplace=True)

                        min_cbm = groupdf.loc[row_pointer,'CBM*Min loading qty']
                        max_qty = groupdf.loc[row_pointer,'Container Quantity (Piece)']
                        min_qty = groupdf.loc[row_pointer,'Min loading qty']
                        
                        curr_qty = groupdf.loc[row_pointer,'Quantity']
                        curr_ctns = groupdf.loc[row_pointer,'TTL CTNS']
                        curr_qty_per_carton = groupdf.loc[row_pointer,'Qty/\nCarton']
                        curr_cbm_per_carton = groupdf.loc[row_pointer,'Width (L) cm']*\
                                                groupdf.loc[row_pointer,'Depth (W) cm']*\
                                                groupdf.loc[row_pointer,'Height (H) cm']/1000000

                        r1=ttl_cbm % min_cbm
                        r2=ttl_cbm % max_cbm

                        if r1!=r2:
                            target_qty=min_qty if min(r1,r2)==r1 else max_qty
                        else:
                            target_qty=max_qty
                        
                        target_ctns = target_qty/curr_qty_per_carton
                        target_cbm = round(target_ctns*curr_cbm_per_carton,3)
                        groupdf.loc[row_pointer,'Quantity']=target_qty
                        groupdf.loc[row_pointer,'TTL CTNS']=target_ctns
                        groupdf.loc[row_pointer,'TTL CBM']=target_cbm

                        groupdf[row_pointer:row_pointer+1][detail_write_columns].to_excel(writer, sheet_name=sheet_name, header=first_group_flag,  index=False,startrow=writer_pointer-1)             
                        
                        item_start_list.append(writer_pointer+1 if first_group_flag else writer_pointer)        
                        writer_pointer+=1
                        if first_group_flag:
                            item_end_list.append(writer_pointer)
                            writer_pointer+=1
                        else:
                            item_end_list.append(writer_pointer-1)
                        
                        first_group_flag=False
                        formula_list.append(writer_pointer)
                        writer_pointer+=1
                        start_list.append(item_start_list[-1])
                        end_list.append(formula_list[-1])

                        new_qty=curr_qty-target_qty
                        new_ctns=curr_ctns-target_ctns
                        new_cbm=round(new_ctns*curr_cbm_per_carton,3)
                        new_df.loc[0,'Quantity']=new_qty
                        new_df.loc[0,'TTL CTNS']=new_ctns
                        new_df.loc[0,'TTL CBM']=new_cbm
                        new_df.loc[0,'split_flag']=new_df.loc[0,'TTL CBM']>=new_df.loc[0,'CBM*Min loading qty']
                        new_df.loc[0,'CLP']=None

                        groupdf.drop(row_pointer,inplace=True)
                        groupdf=pd.concat([groupdf,new_df],axis=0)
                    else:
                        groupdf[row_pointer:row_pointer+1][detail_write_columns].to_excel(writer, sheet_name=sheet_name, header=first_group_flag,  index=False,startrow=writer_pointer-1)             
                        
                        item_start_list.append(writer_pointer if first_group_flag else writer_pointer)
                        writer_pointer+=1
                        if first_group_flag:
                            item_end_list.append(writer_pointer)
                            writer_pointer+=1
                        else:
                            item_end_list.append(writer_pointer-1)
                        
                        first_group_flag=False
                        formula_list.append(writer_pointer)
                        writer_pointer+=1
                        start_list.append(item_start_list[-1])
                        end_list.append(formula_list[-1])
                    
                        groupdf.drop(row_pointer,inplace=True)

                    temp_formula_list_end+=1
                else:
                    break
            

            # -------再写CLP没有计算的部分
            groupdf=groupdf[detail_write_columns]
            ttl_cbm=groupdf['TTL CBM'].sum()

            if ttl_cbm<=66:
                if temp_formula_list_end!=0:
                    temp_formula_list=[formula_list[j] for j in range(-1,-1-temp_formula_list_end,-1)]
                    temp_formula_list.reverse()
                    formula_add_dict[writer_pointer]=temp_formula_list
                    j=1
                    for formula_index in temp_formula_list:
                        worksheet['K'+str(formula_index)]=number_to_ordinal_dict[j]+' TTL:'
                        worksheet['S'+str(formula_index)]='TTL:'
                        j+=1
                    writer_pointer+=1

                # 这时候是上一轮TTL CBM正好在min和max中间，前面drop掉之后这里其实已经没有数据了  改了list就下一个组
                if ttl_cbm==0:
                    end_list[-1]=writer_pointer-1
                    grey_list.append(writer_pointer)
                else:
                    # 这时候是只有CFS或一个40HQ,直接写，记录公式行
                    groupdf['CLP']=splitCLP(ttl_cbm,'').split(',')[0]
                    groupdf.to_excel(writer, sheet_name=sheet_name, header=first_group_flag,  index=False,startrow=writer_pointer-1,)
                    
                    item_start_list.append(writer_pointer+1 if first_group_flag else writer_pointer)                
                    writer_pointer+=len(groupdf)
                    if first_group_flag:
                        item_end_list.append(writer_pointer)
                        writer_pointer+=1
                    else:
                        item_end_list.append(writer_pointer-1)
                    
                    first_group_flag=False
                    formula_list.append(writer_pointer)
                    writer_pointer+=1
                    grey_list.append(writer_pointer)
                    start_list.append(item_start_list[-1])
                    end_list.append(formula_list[-1])
            
            else:
                # 这时候可能有多组，取决于有x个40HQ和1个CFS
                split_result = splitCLP(ttl_cbm,'')
                split_res_list = split_result.split(',')
                split_res_list = list(filter(lambda x:len(x)!=0,split_res_list))

                box_qty=int(split_res_list[0].split('*')[0])
                edge_list=[]
                for i in range(box_qty):
                    edge_list.append(66)
                edge_list.append(0)

                row_pointer = 0
                temp_sum = 0
                split_flag=False
                for i in range(0,len(edge_list)):
                    edge = edge_list[i]
                    if edge!=0:
                        while row_pointer<len(groupdf):
                            current_cbm = groupdf.iloc[row_pointer]['TTL CBM']
                            if temp_sum+current_cbm<edge:
                                row_pointer+=1
                                temp_sum+=current_cbm
                            else:
                                # 计算并填充当前的df
                                curr_width = groupdf.iloc[row_pointer]['Width (L) cm']
                                curr_depth = groupdf.iloc[row_pointer]['Depth (W) cm']
                                curr_height = groupdf.iloc[row_pointer]['Height (H) cm']
                                cbm_diff = edge-temp_sum
                                ori_qty = groupdf.iloc[row_pointer]['Quantity']
                                
                                temp_ctns= int(cbm_diff/(curr_width*curr_depth*curr_height)/1000000)
                                temp_cbm = temp_ctns*(curr_width*curr_depth*curr_height)/1000000
                                while temp_cbm+temp_sum<edge:
                                    temp_ctns+=1
                                    temp_cbm = temp_ctns*(curr_width*curr_depth*curr_height)/1000000
                                
                                # 需要更改的row
                                target_cbm = temp_cbm
                                target_ctns = temp_ctns
                                current_qpc = groupdf.iloc[row_pointer]['Qty/\nCarton']
                                target_qty = temp_ctns * current_qpc

                                new_qty = ori_qty-target_qty
                                new_ctns = int(new_qty/current_qpc)
                                new_cbm = (curr_width*curr_depth*curr_height)/1000000*new_ctns

                                # 存一个备份给下一个df,进行split
                                new_df=pd.DataFrame(groupdf.iloc[row_pointer]).T
                                new_df.reset_index(drop=True,inplace=True)
                                
                                groupdf.loc[row_pointer,'TTL CBM']=target_cbm
                                groupdf.loc[row_pointer,'TTL CTNS']=target_ctns
                                groupdf.loc[row_pointer,'Quantity']=target_qty

                                groupdf.loc[0:row_pointer,'CLP']='1*40HQ'

                                groupdf[0:row_pointer+1].to_excel(writer, sheet_name=sheet_name, header=first_group_flag,  index=False,startrow=writer_pointer-1)
                                
                                item_start_list.append(writer_pointer+1 if first_group_flag else writer_pointer)
                                writer_pointer+=row_pointer
                                if first_group_flag:
                                    writer_pointer+=1
                                item_end_list.append(writer_pointer)
                                writer_pointer+=1
                                formula_list.append(writer_pointer)
                                start_list.append(item_start_list[-1])
                                end_list.append(formula_list[-1])
                                first_group_flag=False

                                writer_pointer+=1

                                # 修改备份df的数据，此时都是CFS
                                new_df.loc[0,'TTL CBM']=new_cbm
                                new_df.loc[0,'TTL CTNS']=new_ctns
                                new_df.loc[0,'Quantity']=new_qty

                                groupdf.drop([x for x in range(0,row_pointer+1)],inplace=True)
                                groupdf=pd.concat([new_df,groupdf],axis=0)
                                groupdf.reset_index(drop=True,inplace=True)

                                split_flag=True
                                row_pointer = 0
                                temp_sum = 0

                                break
                    else:
                        if i>=1 or temp_formula_list_end!=0:
                            temp_formula_list=[formula_list[j] for j in range(-1,-1-i-temp_formula_list_end,-1)]
                            temp_formula_list.reverse()
                            formula_add_dict[writer_pointer]=temp_formula_list
                            j=1
                            for formula_index in temp_formula_list:
                                worksheet['K'+str(formula_index)]=number_to_ordinal_dict[j]+' TTL:'
                                worksheet['S'+str(formula_index)]='TTL:'
                                j+=1
                            writer_pointer+=1
                        
                        groupdf.loc[0:row_pointer,'CLP']='CFS'
                        groupdf[0:].to_excel(writer, sheet_name=sheet_name, header=False, index=False,startrow=writer_pointer-1)
                                
                        item_start_list.append(writer_pointer)
                        writer_pointer+=len(groupdf)
                        item_end_list.append(writer_pointer-1)
                        formula_list.append(writer_pointer)
                        start_list.append(item_start_list[-1])
                        end_list.append(formula_list[-1])

                        writer_pointer+=1
                        grey_list.append(writer_pointer)

        #---------------format---------------------
        border=Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
                
        print('item_start_list:',item_start_list)
        print('item_end_list:',item_end_list)
        print('formula_list:',formula_list)
        print('formula_add_dict:',formula_add_dict)
        print('start_list:',start_list)
        print('end_list:',end_list)
        # split_item_list=list(set(split_item_list))
        # print('split_item_list:',split_item_list)
        print('grey_list:',grey_list)

        worksheet.row_dimensions[1].height=row_height_dict['sheet_title']
        worksheet['F1'].font=Font(name="Calibri", size=28,bold=True)
        # 所有行（表格内容）字体设置
        for i in range(2, writer_pointer+1):
            if i==2:
                for cell in worksheet[i]:
                    cell.font=Font(name="Calibri", size=11,bold=True)
                    cell.alignment = Alignment(wrapText=True,horizontal='center',vertical='center')
                    cell.border=border
            else:
                for cell in worksheet[i]:
                    cell.font=Font(name="Calibri", size=11)
                    cell.alignment = Alignment(horizontal='center',vertical='center')
                    cell.border=border
                    if i in grey_list:
                        cell.fill=PatternFill(fill_type='solid',fgColor='D0CECE')

                if i in grey_list:
                    worksheet.row_dimensions[i].height=row_height_dict['grey_content']   
                else:
                    worksheet.row_dimensions[i].height=row_height_dict['table_content'] 

        # 加和公式和单元格数字格式设置
        for i in range(0,len(item_start_list)):
            if worksheet['X'+str(item_start_list[i])].value=='CFS':
                worksheet['K'+str(formula_list[i])]='CFS TTL:'
                worksheet['S'+str(formula_list[i])]='CFS TTL:'
                for cell in worksheet[formula_list[i]]:
                    cell.font=Font(name="Calibri", size=12,bold=True)
            else:
                worksheet['S'+str(formula_list[i])]='TTL:'

            for char in ['L','T','V','U','W']:
                worksheet[char+str(formula_list[i])]='=SUM('+char+str(item_start_list[i])+':'+char+str(item_end_list[i])+')'
            worksheet['U'+str(formula_list[i])].number_format='0.00'
            worksheet['V'+str(formula_list[i])].number_format='0.00'
            worksheet['W'+str(formula_list[i])].number_format='0.000'

            for r in range(item_start_list[i],item_end_list[i]+1):
                worksheet['M'+str(r)].number_format='m/d/yyyy'
                worksheet['N'+str(r)].number_format='m/d/yyyy'
                worksheet['O'+str(r)].number_format='m/d/yyyy'
                for char in ['U','V','Y']:
                    worksheet[char+str(r)].number_format='0.00'
                worksheet['W'+str(r)].number_format='0.000'
                for char in ['L','T','V','U','W']:
                    worksheet[char+str(r)].font=Font(name="Calibri", size=11,bold=True)
                worksheet['T'+str(r)]='=L'+str(r)+'/P'+str(r)
                worksheet['U'+str(r)]='=T'+str(r)+'*Q'+str(r)
                worksheet['V'+str(r)]='=T'+str(r)+'*R'+str(r)
                worksheet['W'+str(r)]='=round(Z'+str(r)+'*AA'+str(r)+'*Y'+str(r)+'/1000000*T'+str(r)+',3)'

            for char in ['A','B','C','D','X']:
                if char=='X':
                    worksheet.merge_cells(char+str(item_start_list[i])+':'+char+str(formula_list[i]))
                    worksheet[char+str(item_start_list[i])].font = Font(name="Calibri", size=11,color='ff0000',bold=True)
                else:
                    worksheet.merge_cells(char+str(start_list[i])+':'+char+str(end_list[i]))


        # 出现分割时，加和公式以及字体设置
        formula_column_list=['L','T','V','U','W']
        for item in formula_add_dict:
            for col in formula_column_list:
                if col=='L' or col=='T':
                    worksheet[chr(ord(col)-1)+str(item)]='CY TTL:'
                worksheet[col+str(item)]='=SUM('+col+(','+col).join([str(x) for x in formula_add_dict[item]])+')'
            for cell in worksheet[item]:
                cell.font=Font(name="Calibri", size=12,bold=True)

            for row in formula_add_dict[item]:
                for cell in worksheet[row]:
                    cell.font=Font(name="Calibri", size=11,bold=True)
            
            worksheet['X'+str(item)]=str(len(formula_add_dict[item]))+'*40HQ'
            worksheet['X'+str(item)].font = Font(name="Calibri", size=12,color='ff0000',bold=True)
            for char in ['S','T','V','U','W','X']:
                worksheet[char+str(item)].fill = PatternFill("solid", fgColor="ffff00")
        

        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.papersize=9 #A4
        # worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        worksheet.page_setup.scale=USER_CONFIG[user]['SCALE_FOR_DETAIL'] 
        worksheet.page_margins = PageMargins(top=1, bottom=1, left=0.25,right=0.25) 
        
    writer.close()

    print()
    print('---------------------------------------------')
    summary_df=pd.DataFrame()
    for sheet_name in sheet_name_list:
        if sheet_name.find(' 分柜')!=-1:
            print('Labelling battery item in '+sheet_name+'...')
            write_result_df=pd.read_excel(document_dir+document_name,sheet_name=sheet_name,header=1,dtype=str)
            # write_result_df['2nd Item Number']=write_result_df['2nd Item Number'].astype(str)
            write_result_df.dropna(subset=['groupIndex'],inplace=True)
            write_result_df.dropna(subset=['2nd Item Number'],inplace=True)
            write_result_df['groupIndex']=write_result_df['groupIndex'].astype(int)
            write_result_df['Quantity']=write_result_df['Quantity'].astype(int)
            
            
            # ---battery item
            battery_item_index_list=write_result_df[write_result_df['2nd Item Number'].isin(battery_item_list)].index
            battery_item_index_list=[x+3 for x in battery_item_index_list]

            workbook = load_workbook(document_dir+document_name)
            worksheet = workbook[sheet_name]
            if len(battery_item_index_list)!=0:
                for index in battery_item_index_list:
                    worksheet['I'+str(index)].font = Font(name="Calibri", size=11,color='ff0000')
            
            # ---split item
            color_pointer=0
            for index in write_result_df['groupIndex'].unique():
                temp_df=write_result_df[write_result_df['groupIndex']==index]
                temp_df.dropna(how='all',inplace=True)
                split_item_list=temp_df[temp_df['2nd Item Number'].duplicated()]['2nd Item Number'].to_list()

                for item in split_item_list:
                    temp_index_list=temp_df[temp_df['2nd Item Number'].isin([item])].index
                    temp_index_list=[x+3 for x in temp_index_list]

                    color_now=split_item_color_list[color_pointer]
                    for split_item_index in temp_index_list:
                        for col in ['H','I','J','K','L']:
                            worksheet[col+str(split_item_index)].fill = PatternFill("solid", fgColor=color_now)
                    color_pointer=(color_pointer+1)%10    
            
            workbook.save(document_dir+document_name)
            workbook.close()

            # ---get summary data
            for index in write_result_df['groupIndex'].unique():
                # print("index: ",index)
                temp_df=write_result_df[write_result_df['groupIndex']==index]
                clp_result_df=temp_df.groupby('CLP')['CLP'].count().reset_index(name='qty')
                no_cfs_clp_result_df=clp_result_df[clp_result_df['CLP']!='CFS']
                cfs_clp_result_df=clp_result_df[clp_result_df['CLP']=='CFS']

                # print("clp_result_df: ",clp_result_df)
                clp_result_list=[]
                if len(no_cfs_clp_result_df)>0:
                    clp_result_list.append(str(no_cfs_clp_result_df.loc[0,'qty'])+'*40HQ')
                if len(cfs_clp_result_df)>0:
                    clp_result_list.append('CFS')

                temp_df['CLP']=temp_df['CLP'].ffill()
                # print("temp_df: ",temp_df)
                for clp_result in clp_result_list:
                    df_flag='1*40HQ' if clp_result!='CFS' else 'CFS'
                    df=temp_df[temp_df['CLP']==df_flag]
                    df=df.groupby('2nd Item Number')['Quantity'].sum().reset_index(name='qty')
                    df=pd.merge(df,monthly_data[monthly_data['groupIndex']==index])
                    df['Quantity']=df['qty']
                    del df['qty']
                    df['CLP']=clp_result
                    df=df[summary_write_columns]

                    summary_df=pd.concat([summary_df,df])
            

    # ------------------------------------汇总------------------------------------------------------------
    with  pd.ExcelWriter(document_dir+document_name, engine='openpyxl',mode='a',if_sheet_exists='overlay') as writer:
        for sheet_name in sheet_name_list:
            if sheet_name.find(' 分柜')!=-1:
                continue
            print('---------------------------------------------')
            print(sheet_name)
            
            worksheet = writer.sheets[sheet_name]
            
            for key,value in column_width_dict.items():
                worksheet.column_dimensions[key].width=value
            worksheet.column_dimensions['Z'].width=10
            worksheet.column_dimensions['AA'].width=12
            
            worksheet.column_dimensions['AB'].hidden=1
            worksheet.column_dimensions['AC'].hidden=1
            worksheet.column_dimensions['AD'].hidden=1
            worksheet.column_dimensions['AE'].hidden=1
            
            #----------------------content----------------------
            writer_pointer=0
            start_list=[]
            end_list=[]
            item_start_list=[]
            item_end_list=[]
            formula_list=[]
            formula_add_dict={}
            grey_list=[]

            df=summary_df[summary_df['Vendor Name']==sheet_name]
            worksheet['F1']='TGT - '+month_dict[df['First Ship Date'].sort_values().tolist()[0].month]+' - '+sheet_name
            writer_pointer+=1

            first_group_flag=False
            for index in df['groupIndex'].unique():
                if list(df['groupIndex'].unique()).index(index)==0:
                    first_group_flag=True
                writer_pointer+=1
                groupdf=df[df['groupIndex']==index]

                for clp_result in groupdf['CLP'].unique():
                    clp_groupdf=groupdf[groupdf['CLP']==clp_result]
                    clp_groupdf.to_excel(writer, sheet_name=sheet_name, header=first_group_flag,  index=False,startrow=writer_pointer-1)             

                    item_start_list.append(writer_pointer+1 if first_group_flag else writer_pointer)        
                    writer_pointer+=len(clp_groupdf)
                    if first_group_flag:
                        item_end_list.append(writer_pointer)
                        writer_pointer+=1
                    else:
                        item_end_list.append(writer_pointer-1)
                    
                    first_group_flag=False
                    formula_list.append(writer_pointer)
                    writer_pointer+=1
                    start_list.append(item_start_list[-1])
                    end_list.append(formula_list[-1])

                grey_list.append(writer_pointer)    
            

            #---------------format---------------------
            border=Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
            
            print('item_start_list:',item_start_list)
            print('item_end_list:',item_end_list)
            print('formula_list:',formula_list)
            print('formula_add_dict:',formula_add_dict)
            print('start_list:',start_list)
            print('end_list:',end_list)
            print('grey_list:',grey_list)

            worksheet.row_dimensions[1].height=row_height_dict['sheet_title']
            worksheet['F1'].font=Font(name="Calibri", size=28,bold=True)

            # 所有行（表格内容）字体设置
            for i in range(2, writer_pointer+1):
                if i==2:
                    for cell in worksheet[i]:
                        cell.font=Font(name="Calibri", size=11,bold=True)
                        cell.alignment = Alignment(wrapText=True,horizontal='center',vertical='center')
                        cell.border=border
                else:
                    for cell in worksheet[i]:
                        cell.font=Font(name="Calibri", size=11)
                        cell.alignment = Alignment(horizontal='center',vertical='center')
                        cell.border=border
                        if i in grey_list:
                            cell.fill=PatternFill(fill_type='solid',fgColor='D0CECE')

                    if i in grey_list:
                        worksheet.row_dimensions[i].height=row_height_dict['grey_content']   
                    else:
                        worksheet.row_dimensions[i].height=row_height_dict['table_content'] 

            # 加和公式和单元格数字格式设置
            for i in range(0,len(item_start_list)):
                if worksheet['X'+str(item_start_list[i])].value=='CFS':
                    worksheet['K'+str(formula_list[i])]='CFS TTL:'
                    worksheet['S'+str(formula_list[i])]='CFS TTL:'
                else:
                    worksheet['K'+str(formula_list[i])]='CY TTL:'
                    worksheet['S'+str(formula_list[i])]='CY TTL:'
                    worksheet['X'+str(item_start_list[i])].fill = PatternFill("solid", fgColor='FFFF00')
                    worksheet['W'+str(formula_list[i])].fill = PatternFill("solid", fgColor='FFFF00')
                for cell in worksheet[formula_list[i]]:
                    cell.font=Font(name="Calibri", size=12,bold=True)
                

                for char in ['L','T','V','U','W','Z']:
                    worksheet[char+str(formula_list[i])]='=SUM('+char+str(item_start_list[i])+':'+char+str(item_end_list[i])+')'
                worksheet['U'+str(formula_list[i])].number_format='0.00'
                worksheet['V'+str(formula_list[i])].number_format='0.00'
                worksheet['W'+str(formula_list[i])].number_format='0.000'

                for r in range(item_start_list[i],item_end_list[i]+1):
                    worksheet['M'+str(r)].number_format='m/d/yyyy'
                    worksheet['N'+str(r)].number_format='m/d/yyyy'
                    worksheet['O'+str(r)].number_format='m/d/yyyy'
                    for char in ['U','V','Y']:
                        worksheet[char+str(r)].number_format='0.00'
                    worksheet['W'+str(r)].number_format='0.000'
                    for char in ['L','T','V','U','W']:
                        worksheet[char+str(r)].font=Font(name="Calibri", size=11,bold=True)
                    worksheet['T'+str(r)]='=L'+str(r)+'/P'+str(r)
                    worksheet['U'+str(r)]='=T'+str(r)+'*Q'+str(r)
                    worksheet['V'+str(r)]='=T'+str(r)+'*R'+str(r)
                    worksheet['Z'+str(r)]='=L'+str(r)+'*Y'+str(r)
                    worksheet['W'+str(r)]='=round(AB'+str(r)+'*AD'+str(r)+'*AC'+str(r)+'/1000000*T'+str(r)+',3)'

                for char in ['A','B','C','D','X']:
                    if char=='X':
                        worksheet.merge_cells(char+str(item_start_list[i])+':'+char+str(formula_list[i]))
                        worksheet[char+str(item_start_list[i])].font = Font(name="Calibri", size=11,color='ff0000',bold=True)
                    else:
                        worksheet.merge_cells(char+str(start_list[i])+':'+char+str(end_list[i]))

            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.paperSize=9 #A4
            # worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.page_setup.scale=USER_CONFIG[user]['SCALE_FOR_SUMMARY'] 
            worksheet.page_margins = PageMargins(top=1, bottom=1, left=0.25,right=0.25) 

    print()
    print('---------------------------------------------')   
    for sheet_name in sheet_name_list:
        if sheet_name.find(' 分柜')==-1:
            print('Labelling battery item in '+sheet_name+'...')
            write_result_df=pd.read_excel(document_dir+document_name,sheet_name=sheet_name,header=1,)
            
            write_result_df.dropna(subset=['groupIndex'],inplace=True)
            # ---battery item
            battery_item_index_list=write_result_df[write_result_df['2nd Item Number'].isin(battery_item_list)].index
            battery_item_index_list=[x+3 for x in battery_item_index_list]

            workbook = load_workbook(document_dir+document_name)
            worksheet = workbook[sheet_name]
            if len(battery_item_index_list)!=0:
                for index in battery_item_index_list:
                    worksheet['I'+str(index)].font = Font(name="Calibri", size=11,color='ff0000')
            
            # ---split item
            color_pointer=0
            for index in write_result_df['groupIndex'].unique():
                temp_df=write_result_df[write_result_df['groupIndex']==index]
                temp_df.dropna(how='all',inplace=True)
                split_item_list=temp_df[temp_df['2nd Item Number'].duplicated()]['2nd Item Number'].to_list()

                for item in split_item_list:
                    temp_index_list=temp_df[temp_df['2nd Item Number'].isin([item])].index
                    temp_index_list=[x+3 for x in temp_index_list]

                    color_now=split_item_color_list[color_pointer]
                    for split_item_index in temp_index_list:
                        for col in ['H','I','J','K','L']:
                            worksheet[col+str(split_item_index)].fill = PatternFill("solid", fgColor=color_now)
                    color_pointer=(color_pointer+1)%10    
            
            workbook.save(document_dir+document_name)
            workbook.close()


if __name__ == "__main__":
    root = tk.Tk()
    root.wm_attributes("-topmost", 1)
    root.withdraw()
    messagebox.showinfo("提示", "Starting...")

    folder_path = os.getcwd()+'\\'
    # folder_path = os.getcwd()+'\\CLP\\TGT CLP\\'
    # print(os.getcwd())

    user='Rainbow'
    # user='Lyn'
    print('CURRENT USER: '+user)
    data_file_name = "TGT COVER.xlsx"
    monthly_data, battery_item_list, sheet_name_list = data_extraction(
        file_dir=folder_path, data_file_name=data_file_name
    )
    document_dir = folder_path + "DOCUMENTS\\"
    write_excel(
        monthly_data,
        battery_item_list,
        sheet_name_list,
        document_dir=document_dir,
    )
    messagebox.showinfo("提示", "Completed.")

