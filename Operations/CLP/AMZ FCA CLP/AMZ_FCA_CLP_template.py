import pandas as pd
from tkinter import messagebox
import tkinter as tk
import os
import math
from openpyxl.styles import Font,Alignment, colors,Border,Side,PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl import load_workbook
from datetime import datetime,timedelta
import warnings
from pandas.errors import SettingWithCopyWarning

warnings.simplefilter(action='ignore', category=(SettingWithCopyWarning))

# update log:
# 2025.3.6: 对于battery_item标红，新增master item表种description里包含bubble的两个model：609BZ/617z

CLP_DICT={
    '<25':'CFS',
    '<56':'20GP',
    '<61':'40GP',
    '>=61':'40HQ'}

EDGE_CLP={
    'CFS':0,
    '20GP':25,
    '40GP':56,
    '40HQ':61
}

def splitCLP(ttl_cbm,temp_res):
    if ttl_cbm>=61:
        temp_res+=CLP_DICT['>=61']
        ttl_cbm-=61
    elif ttl_cbm<61 and ttl_cbm>=56:
        temp_res+=CLP_DICT['<61']
        ttl_cbm-=56
    elif ttl_cbm<56 and ttl_cbm>=25:
        temp_res+=CLP_DICT['<56']
        ttl_cbm-=25
    else:
        temp_res+=CLP_DICT['<25']
        ttl_cbm-=25

    
    temp_res+=','
    if ttl_cbm>0:
        temp_res = splitCLP(ttl_cbm,temp_res)
    return temp_res 


def data_extraction(file_dir, data_file_name):
    # ------------------------------------------------------------monthly_data------------------------------------------------------------
    monthly_data=pd.read_excel(file_dir+data_file_name,sheet_name='Monthly Data',dtype=str)
    monthly_data=monthly_data.sort_values(['Vendor Name','Vendor code','FC','First Ship Date','Last Ship Date'])
    monthly_data.reset_index(drop=True,inplace=True)
    monthly_data['Last Ship Date']=monthly_data['Last Ship Date'].apply(lambda x:datetime.strptime(x,'%Y-%m-%d %H:%M:%S'))
    monthly_data['First Ship Date']=monthly_data['First Ship Date'].apply(lambda x:datetime.strptime(x,'%Y-%m-%d %H:%M:%S'))
    monthly_data['Cargo Ready Date']=monthly_data['Cargo Ready Date'].apply(lambda x:datetime.strptime(x,'%Y-%m-%d %H:%M:%S'))

    change_column=['Quantity','Qty/\nCarton', 'Net Weight (kg)','Gross Weight (kg)', 'Cubic\nMeters (per carton)', 'TTL CTNS','TTL NW (KG)', 'TTL GW (KG)', 'TTL CBM']
    for column in change_column:
        monthly_data[column]=monthly_data[column].astype(float)
    
    monthly_data['TTL CBM'] = monthly_data['TTL CBM'].apply(lambda x:round(x,3))

    # -----获取target，遍历
    groupBy = monthly_data.groupby(['Vendor Name','Vendor code','FC',])
    groupIndex=1

    for name,group in groupBy:
        # print('-------------------')
        # print(name)
        temp = groupBy.get_group(name)
        if len(temp)>1:
            # print('BEFORE')
            # print(temp[['First Ship Date','Last Ship Date']],)
            
            # 判断时间窗口是否重叠：Sa≥Eb | Sb≥Ea，则不重叠；反之重叠。
            target_first_ship_date=temp.iloc[0]['First Ship Date']
            target_last_ship_date=temp.iloc[0]['Last Ship Date']

            inner_group_index=1
            start_flag=True
            for index,rows in temp.iterrows():
                if start_flag:        
                    temp.loc[index,'overlap']=inner_group_index
                    start_flag=False
                    continue
                
                if target_first_ship_date<rows['Last Ship Date']+timedelta(days=-1) and target_last_ship_date>rows['First Ship Date']+timedelta(days=1):
                    temp.loc[index,'overlap']=inner_group_index
                else:
                    target_first_ship_date=rows['First Ship Date']
                    target_last_ship_date=rows['Last Ship Date']
                    inner_group_index+=1
                    temp.loc[index,'overlap']=inner_group_index
                    
            # print('AFTER')
            # print(temp[['First Ship Date','Last Ship Date','overlap']])

            overlap_list=temp['overlap'].drop_duplicates().tolist()
            for flag in overlap_list:
                monthly_data.loc[temp[temp['overlap']==flag].index,'groupIndex']=groupIndex
                groupIndex+=1
                # print(monthly_data.loc[temp[temp['overlap']==flag].index][['Vendor Name','Vendor code','FC','First Ship Date','Last Ship Date','groupIndex']])
                # print()
        else:
            monthly_data.loc[temp.index,'groupIndex']=groupIndex
            groupIndex+=1
            # print(monthly_data.loc[temp.index][['Vendor Name','Vendor code','FC','First Ship Date','Last Ship Date','groupIndex']])
            # print()

    
    # ------------------------------------------------------------battery_item------------------------------------------------------------
    battery_item=monthly_data[monthly_data['Battery']=='Yes']['2nd Item Number'].drop_duplicates().tolist()
    special_item_list=['609BZ','609BAZ','617Z','617CZ','617AZ','617AU']
    for item in special_item_list:
        battery_item.append(item)

    # ------------------------------------------------------------asin_fc------------------------------------------------------------
    asin_fc=pd.read_excel(file_dir+data_file_name,sheet_name='Config',dtype=str)
    asin_fc=asin_fc['ASIN FC'].tolist()


    return monthly_data,battery_item,asin_fc


def write_excel(monthly_data, battery_item,asin_fc,document_dir):
    for vendor in monthly_data['Vendor Name'].unique():
        print('---------------------------------------------')
        print(vendor)
        temp_df=monthly_data[monthly_data['Vendor Name']==vendor]
        split_item_df=pd.DataFrame()

        document_name=str(datetime.now().month).zfill(2)+str(datetime.now().day).zfill(2)+'_'+vendor.strip()+'.xlsx'
        writer = pd.ExcelWriter(document_dir+document_name, engine='openpyxl')
        pd.DataFrame().to_excel(writer, sheet_name='Sheet1', index=True)
        worksheet = writer.sheets['Sheet1']
        worksheet.column_dimensions['B'].width=10
        worksheet.column_dimensions['E'].width=12
        worksheet.column_dimensions['F'].width=12
        worksheet.column_dimensions['H'].width=14
        worksheet.column_dimensions['J'].width=11
        worksheet.column_dimensions['K'].width=11
        worksheet.column_dimensions['L'].width=11
        worksheet.column_dimensions['R'].width=10
        worksheet.column_dimensions['S'].width=10
        
        #-----------------------------------------------------content-----------------------------------------------------
        writer_pointer=1
        header_list=[]
        end_list=[]
        item_start_list=[]
        item_end_list=[]
        formula_list=[]
        formula_add_dict={}
        split_item_list=[]
        # ----先写不需要拆分的df
        df1=temp_df[temp_df['FC'].isin(asin_fc)]
        for index in df1['groupIndex'].unique():
            groupdf=df1[df1['groupIndex']==index]
            
            worksheet['A'+str(writer_pointer)]='AMU840N'
            worksheet['G'+str(writer_pointer)]='ECDD: '
            worksheet['J'+str(writer_pointer)]='vdr# '
            worksheet['K'+str(writer_pointer)]=groupdf['Vendor code'].unique()[0]
            
            if len(header_list)!=0:
                end_list.append(writer_pointer-2)
            header_list.append(writer_pointer)
            
            writer_pointer+=1
            groupdf['Vendor code']='CFS'
            groupdf.rename(columns={'Vendor code':'CLP','FC':'DC#'},inplace=True)
            groupdf.reset_index(drop=True,inplace=True)
            
            del groupdf['Battery'],groupdf['groupIndex'],groupdf['Ship To Description']
            groupdf = groupdf.sort_values(['Customer PO','ASIN# or SKU#... '])
            groupdf.to_excel(writer, sheet_name='Sheet1', index=False, startrow=writer_pointer-1)
        
            item_start_list.append(writer_pointer+1)
            writer_pointer+=len(groupdf)
            item_end_list.append(writer_pointer)
            writer_pointer+=1
            
            formula_list.append(writer_pointer)
            
            writer_pointer+=2
        
        # ----再写可能需要拆分的df
        df2=temp_df[~temp_df['FC'].isin(asin_fc)]
        for index in df2['groupIndex'].unique():
            groupdf=df2[df2['groupIndex']==index]
            
            vendor_code=groupdf['Vendor code'].unique()[0]
            worksheet['A'+str(writer_pointer)]='AMU840N'
            worksheet['G'+str(writer_pointer)]='ECDD: '
            worksheet['J'+str(writer_pointer)]='vdr# '
            worksheet['K'+str(writer_pointer)]=vendor_code
        
            if len(header_list)!=0:
                end_list.append(writer_pointer-2)
            header_list.append(writer_pointer)
            
            writer_pointer+=1
        
            groupdf=groupdf.sort_values(['TTL CBM'],ascending=False)
            groupdf.reset_index(drop=True,inplace=True)
            # 先填充CFS
            groupdf['Vendor code']='CFS'
            groupdf.rename(columns={'Vendor code':'CLP','FC':'DC#'},inplace=True)
        
            del groupdf['Battery'],groupdf['groupIndex'],groupdf['Ship To Description']
            
            ttl_cbm=groupdf['TTL CBM'].sum()
        
            # 如果TTL CBM<=80再计算，否则不计算
            if ttl_cbm<=80:
                split_result = splitCLP(ttl_cbm,'')
                split_res_list = split_result.split(',')
                split_res_list = list(filter(lambda x:len(x)!=0,split_res_list))
                edge_list=[EDGE_CLP[res] for res in split_res_list]
                
                row_pointer = 0 
                temp_sum=0
                split_flag=False
                for edge in edge_list:
                    # print('edge：',edge)
                    # print('split_item_list',split_item_list)

                    # !!Section:目标edge=0，row_pointer还没开始迭代，可能是不用split的，直接to_excel[249]，但也可能是被split过，则需要记一下行号
                    if edge==0:
                        groupdf = groupdf.sort_values(['Customer PO','ASIN# or SKU#... '])
                        groupdf.reset_index(inplace=True,drop=True)

                        if split_flag:
                            # print('edge==0, split_flag is True',writer_pointer)
                            writer_pointer+=1
                            worksheet['A'+str(writer_pointer)]='AMU840N'
                            worksheet['G'+str(writer_pointer)]='ECDD: '
                            worksheet['J'+str(writer_pointer)]='vdr# '
                            worksheet['K'+str(writer_pointer)]=vendor_code
                            if len(header_list)!=0:
                                end_list.append(writer_pointer-2)
                            header_list.append(writer_pointer)
                            
                            writer_pointer+=1
                            
                            split_item_index = groupdf[(groupdf['Customer PO']==split_item_PO) & (groupdf['ASIN# or SKU#... ']==split_item_ASIN)].index[0]
                            # print('split_item_index:',split_item_index)
                            split_item_list.append(writer_pointer+split_item_index+1)

                        
                        groupdf.to_excel(writer, sheet_name='Sheet1', index=False, startrow=writer_pointer-1)
                
                        item_start_list.append(writer_pointer+1)
                        writer_pointer+=len(groupdf)
                        item_end_list.append(writer_pointer)
                        writer_pointer+=1
                        formula_list.append(writer_pointer)
            
                        writer_pointer+=2
                        break
                    
                    start_row = row_pointer
                    write_flag = False
                    
                    while row_pointer<len(groupdf):
                        # print(edge,"----",row_pointer)
                        current_cbm = groupdf.iloc[row_pointer]['TTL CBM']
                        if temp_sum+current_cbm<edge:
                            row_pointer+=1
                            temp_sum+=current_cbm
                        elif temp_sum+current_cbm == edge:
                            write_flag = True
                            break
                        else:
                            # 计算并填充当前的df
                            target_cbm = edge-temp_sum
                            current_unit_cbm = groupdf.iloc[row_pointer]['Cubic\nMeters (per carton)']
                            target_ctns = math.ceil(target_cbm/current_unit_cbm)
                            current_qty = groupdf.iloc[row_pointer]['Qty/\nCarton']
                            target_qty = target_ctns * current_qty
        
                            # 存一个备份给下一个df,进行split
                            new_df=pd.DataFrame(groupdf.iloc[row_pointer]).T
                            new_df.reset_index(drop=True,inplace=True)
                            # print('new_df',new_df)
                            
                            groupdf.loc[row_pointer,'TTL CBM']=target_cbm
                            groupdf.loc[row_pointer,'TTL CTNS']=target_ctns
                            groupdf.loc[row_pointer,'Quantity']=target_qty
        
                            groupdf.loc[start_row:row_pointer,'CLP']=split_res_list[edge_list.index(edge)]
                            # print('writing groupdf:',groupdf)

                            # !!Section:row_pointer到这一行发现加起来又超过了edge，需要进行split，且这个frame有一行已经被spilt过
                            if split_flag:
                                # 先记下来此次需要被split的PO和ASIN
                                new_split_item_PO = groupdf.loc[row_pointer,'Customer PO']
                                new_split_item_ASIN = groupdf.loc[row_pointer,'ASIN# or SKU#... ']
                                
                                # ---排序后记下上次被split的index,和本次被Split的index
                                groupdf[start_row:row_pointer+1] = groupdf[start_row:row_pointer+1].sort_values(['Customer PO','ASIN# or SKU#... '])
                                groupdf[start_row:row_pointer+1].reset_index(drop=True,inplace=True)
                                
                                split_item_index = groupdf[start_row:row_pointer+1][(groupdf[start_row:row_pointer+1]['Customer PO']==split_item_PO) & (groupdf[start_row:row_pointer+1]['ASIN# or SKU#... ']==split_item_ASIN)].index[0]
                                split_item_list.append(writer_pointer+split_item_index)
                                
                                split_item_ASIN = new_split_item_ASIN
                                split_item_PO = new_split_item_PO
                                split_item_index = groupdf[start_row:row_pointer+1][(groupdf[start_row:row_pointer+1]['Customer PO']==split_item_PO) & (groupdf[start_row:row_pointer+1]['ASIN# or SKU#... ']==split_item_ASIN)].index[0]
                                split_item_list.append(writer_pointer+split_item_index)
                                
                                groupdf[start_row:row_pointer+1].to_excel(writer, sheet_name='Sheet1', index=False, startrow=writer_pointer-1,header=False)
    
                                item_start_list.append(writer_pointer)
                                writer_pointer+=len(groupdf[start_row:row_pointer+1])
                                item_end_list.append(writer_pointer-1)
                                formula_list.append(writer_pointer)
        
                                # 两个非CFS的要拼接公式
                                writer_pointer+=1
                                formula_add_dict[writer_pointer]=[formula_list[-2],formula_list[-1]]

                            # !!Section:row_pointer到这一行发现加起来超过了edge，需要进行split，但这个frame还没有被spilt过，则是这一单的开头frame
                            else:
                                # 先记下来被split的PO和asin，以便下次可能会用；同时记下本次行号，存入split_item_list
                                split_item_PO = groupdf.loc[row_pointer,'Customer PO']
                                split_item_ASIN = groupdf.loc[row_pointer,'ASIN# or SKU#... ']

                                groupdf[start_row:row_pointer+1] = groupdf[start_row:row_pointer+1].sort_values(['Customer PO','ASIN# or SKU#... '])
                                groupdf[start_row:row_pointer+1].reset_index(drop=True,inplace=True)

                                groupdf[start_row:row_pointer+1].to_excel(writer, sheet_name='Sheet1', index=False, startrow=writer_pointer-1)
                                split_flag=True 
        
                                split_item_index = groupdf[start_row:row_pointer+1][(groupdf[start_row:row_pointer+1]['Customer PO']==split_item_PO) & (groupdf[start_row:row_pointer+1]['ASIN# or SKU#... ']==split_item_ASIN)].index[0]
                                split_item_list.append(writer_pointer+split_item_index+1)
        
                                item_start_list.append(writer_pointer+1)
                                writer_pointer+=len(groupdf[start_row:row_pointer+1])
                                item_end_list.append(writer_pointer)
                                writer_pointer+=1
                                formula_list.append(writer_pointer)
                
                            writer_pointer+=1
        
                            # 修改备份df的数据
                            new_cbm = current_cbm-target_cbm
                            new_ctns = math.ceil(new_cbm/current_unit_cbm)
                            new_qty = new_ctns * current_qty
        
                            new_df.loc[0,'TTL CBM']=new_cbm
                            new_df.loc[0,'TTL CTNS']=new_ctns
                            new_df.loc[0,'Quantity']=new_qty
        
                            groupdf.drop([x for x in range(start_row,row_pointer+1)],inplace=True)
                            groupdf=pd.concat([new_df,groupdf],axis=0)
                            groupdf.reset_index(drop=True,inplace=True)
                            # print(groupdf)
                            
                            row_pointer=0
                            temp_sum=0
                            break
        
                    # !!Section:到这里是因为目标edge不为0，但row_pointer又刚好迭代到TTL CBM刚刚好加起来等于edge
                    if write_flag:
                        # print('in the write_flag, split_flag:',split_flag,writer_pointer)
                        groupdf.loc[start_row:row_pointer,'CLP']=split_res_list[edge_list.index(edge)]

                        # !!Section:被split过，所以记行号存起来
                        if split_flag:
                            split_item_index = groupdf[start_row:row_pointer+1][(groupdf[start_row:row_pointer+1]['Customer PO']==split_item_PO) & (groupdf[start_row:row_pointer+1]['ASIN# or SKU#... ']==split_item_ASIN)].index[0]
                            split_item_list.append(writer_pointer+split_item_index+1)

                            groupdf[start_row:row_pointer+1] = groupdf[start_row:row_pointer+1].sort_values(['Customer PO','ASIN# or SKU#... '])
                            groupdf[start_row:row_pointer+1].to_excel(writer, sheet_name='Sheet1', index=False, startrow=writer_pointer-1,header=False)
        
                            item_start_list.append(writer_pointer)
                            writer_pointer+=len(groupdf[start_row:row_pointer+1])
                            item_end_list.append(writer_pointer-1)
                            formula_list.append(writer_pointer)

                        # !!Section:没有被split过，所以不用做什么
                        else:
                            groupdf[start_row:row_pointer+1] = groupdf[start_row:row_pointer+1].sort_values(['Customer PO','ASIN# or SKU#... '])
                            groupdf[start_row:row_pointer+1].to_excel(writer, sheet_name='Sheet1', index=False, startrow=writer_pointer-1)
                            
                            groupdf.drop([x for x in range(start_row,row_pointer+1)],inplace=True)
                            groupdf.reset_index(drop=True,inplace=True)
                            split_flag=True # 主要是为了新写下一个frame
        
                            # ...包括sum公式填充
                            item_start_list.append(writer_pointer+1)
                            writer_pointer+=len(groupdf[start_row:row_pointer+1])
                            item_end_list.append(writer_pointer)
                            writer_pointer+=1
                            formula_list.append(writer_pointer)
            
                        writer_pointer+=1
            else:
                groupdf.to_excel(writer, sheet_name='Sheet1', index=False, startrow=writer_pointer-1)
                
                item_start_list.append(writer_pointer+1)
                writer_pointer+=len(groupdf)
                item_end_list.append(writer_pointer)
                writer_pointer+=1
                formula_list.append(writer_pointer)
    
                writer_pointer+=2
        end_list.append(writer_pointer-2)


        #-----------------------------------------------------format-----------------------------------------------------
        border=Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
                
        print('item_start_list:',item_start_list)
        print('item_end_list:',item_end_list)
        print('formula_list:',formula_list)
        print('formula_add_dict:',formula_add_dict)
        print('header_list:',header_list)
        print('end_list:',end_list)
        print('split_item_list:',split_item_list)

        # 所有行（表格内容）字体设置
        for i in range(1, writer_pointer+1):
            for cell in worksheet[i]:
                cell.font=Font(name="Arail", size=11)

        # 加和公式和单元格数字格式设置
        for i in range(0,len(item_start_list)):
            worksheet['I'+str(formula_list[i])]='=SUM(I'+str(item_start_list[i])+':I'+str(item_end_list[i])+')'
            worksheet['Q'+str(formula_list[i])]='=SUM(Q'+str(item_start_list[i])+':Q'+str(item_end_list[i])+')'
            worksheet['R'+str(formula_list[i])]='=SUM(R'+str(item_start_list[i])+':R'+str(item_end_list[i])+')'
            worksheet['S'+str(formula_list[i])]='=SUM(S'+str(item_start_list[i])+':S'+str(item_end_list[i])+')'
            worksheet['T'+str(formula_list[i])]='=SUM(T'+str(item_start_list[i])+':T'+str(item_end_list[i])+')'
                
            for r in range(item_start_list[i],item_end_list[i]+1):
                worksheet['J'+str(r)].number_format='m/d/yyyy'
                worksheet['K'+str(r)].number_format='m/d/yyyy'
                worksheet['L'+str(r)].number_format='m/d/yyyy'
                worksheet['R'+str(r)].number_format='0.00'
                worksheet['S'+str(r)].number_format='0.00'
                worksheet['T'+str(r)].number_format='0.000'

                worksheet['Q'+str(r)]='=I'+str(r)+'/M'+str(r)
                worksheet['R'+str(r)]='=Q'+str(r)+'*N'+str(r)
                worksheet['S'+str(r)]='=Q'+str(r)+'*O'+str(r)
                worksheet['T'+str(r)]='=Q'+str(r)+'*P'+str(r)

            worksheet['R'+str(formula_list[i])].number_format='0.00'
            worksheet['S'+str(formula_list[i])].number_format='0.00'
            worksheet['T'+str(formula_list[i])].number_format='0.000'
            worksheet.merge_cells("U"+str(item_start_list[i])+':U'+str(item_end_list[i]))
            worksheet['U'+str(item_start_list[i])].alignment = Alignment(horizontal='center',vertical='center')
            worksheet['U'+str(item_start_list[i])].font = Font(name="Arail", size=11,color='ff0000')

        # 表头字体以及行高设置
        for index in header_list:
            for cell in worksheet[index]:
                cell.font=Font(name="Arail", size=16,bold=True)
            worksheet.row_dimensions[index].height=50
            for cell in worksheet[index+1]:
                cell.font=Font(name="Arail", size=9,bold=True)
                cell.alignment = Alignment(wrapText=True,horizontal='center',vertical='center')

        # 表格边框设置
        for i in range(0,len(header_list)):
            header_index=header_list[i]+1
            end_index=end_list[i]
            for row in worksheet.iter_rows(min_row=header_index,max_row=end_index):
                for cell in row:
                    cell.border=border

        # 公式字体设置
        for index in formula_list:
            for cell in worksheet[index]:
                cell.font=Font(name="Arail", size=12,bold=True)

        # 出现分割时，加和公式以及字体设置
        for item in formula_add_dict:
            worksheet['I'+str(item)]='=SUM(I'+str(formula_add_dict[item][0])+',I'+str(formula_add_dict[item][1])+')'
            worksheet['Q'+str(item)]='=SUM(Q'+str(formula_add_dict[item][0])+',Q'+str(formula_add_dict[item][1])+')'
            worksheet['R'+str(item)]='=SUM(R'+str(formula_add_dict[item][0])+',R'+str(formula_add_dict[item][1])+')'
            worksheet['S'+str(item)]='=SUM(S'+str(formula_add_dict[item][0])+',S'+str(formula_add_dict[item][1])+')'
            worksheet['T'+str(item)]='=SUM(T'+str(formula_add_dict[item][0])+',T'+str(formula_add_dict[item][1])+')'
            for cell in worksheet[item]:
                cell.font=Font(name="Arail", size=10,bold=True)
        
        # 出现分割时，相关行的颜色填充设置
        if len(split_item_list)!=0:
            for index in split_item_list:
                worksheet['F'+str(index)].fill = PatternFill("solid", fgColor="DDEBF7")
                worksheet['G'+str(index)].fill = PatternFill("solid", fgColor="DDEBF7")
                worksheet['H'+str(index)].fill = PatternFill("solid", fgColor="DDEBF7")
                worksheet['I'+str(index)].fill = PatternFill("solid", fgColor="DDEBF7")

                row_index=len(split_item_df)
                split_item_df.loc[row_index,'Vendor Name']=worksheet['D'+str(index)].value
                split_item_df.loc[row_index,'Customer PO']=worksheet['F'+str(index)].value
                split_item_df.loc[row_index,'2nd Item Number']=worksheet['G'+str(index)].value
                split_item_df.loc[row_index,'Split Qty']=worksheet['I'+str(index)].value

        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins = PageMargins(top=1, bottom=1, left=0.5,right=0.5) 

        # --- Variance record
        if len(split_item_df)!=0:
            split_item_df.drop_duplicates(inplace=True)
            split_item_df=pd.DataFrame(split_item_df.groupby([ 'Vendor Name','Customer PO', '2nd Item Number',])['Split Qty'].sum())
            split_item_df.reset_index(inplace=True)
            split_item_df.rename(columns={'Split Qty':'Validated Quantity'},inplace=True)
            
            variance_df=pd.merge(split_item_df,df2)
            variance_df=variance_df[['Order Number', 'Related Order Number','Vendor Name', 'Sold To Name',
                                    'Customer PO', '2nd Item Number', 'ASIN# or SKU#... ', 'Quantity','Validated Quantity']]
            variance_df=variance_df[variance_df['Validated Quantity']!=variance_df['Quantity']]
            if len(variance_df)!=0:
                variance_df.to_excel(writer, sheet_name='Difference', index=False)
                worksheet = writer.sheets['Difference']
                worksheet.column_dimensions["A"].width = 10
                worksheet.column_dimensions["C"].width = 49
                worksheet.column_dimensions["D"].width = 27
                worksheet.column_dimensions["E"].width = 14
                worksheet.column_dimensions["G"].width = 14
                worksheet.column_dimensions["I"].width = 9.5

                for i in range(1, len(variance_df) + 2):
                    if i==1:
                        for cell in worksheet[i]:
                            cell.font = Font(name="Arail", size=10,bold=True)
                            cell.alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                    else:
                        for cell in worksheet[i]:
                            cell.font = Font(name="Arail", size=11)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = border
                            if cell.col_idx==9:
                                cell.fill = PatternFill("solid", fgColor="FFFF00")

        writer.close()

        # ----为了标注battery item，写完重新读
        print('Labelling battery item in '+vendor+'...')
        write_result_df=pd.read_excel(document_dir+document_name)
        battery_item_index_list=write_result_df[write_result_df['ECDD: '].isin(battery_item)].index
        battery_item_index_list=[x+2 for x in battery_item_index_list]
        
        if len(battery_item_index_list)!=0:
            workbook = load_workbook(document_dir+document_name)
            worksheet = workbook['Sheet1']
            for index in battery_item_index_list:
                worksheet['G'+str(index)].font = Font(name="Arail", size=11,color='ff0000')
            workbook.save(document_dir+document_name)


if __name__ == '__main__':
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    messagebox.showinfo("提示", "Starting...")

    folder_path = os.getcwd()
    data_file_name='AMZ EU Cover.xlsx'
    monthly_data,battery_item,asin_fc = data_extraction(file_dir=folder_path, data_file_name='\\'+data_file_name)
    document_dir=folder_path+'\\DOCUMENTS\\'
    write_excel(monthly_data,battery_item,asin_fc,document_dir=document_dir)
    messagebox.showinfo("提示", "Completed.")
