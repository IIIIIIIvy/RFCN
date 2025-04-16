import pandas as pd
from tkinter import messagebox
import tkinter as tk
import os
import openpyxl
import datetime
from pathlib import Path
from openpyxl.styles import Font,Alignment,Side, Border,PatternFill,colors


def excel_column_formatter(index):
    dic={
        0: '',
        1: 'A',
        2: 'B',
        3: 'C',
        4: 'D',
        5: 'E',
        6: 'F',
        7: 'G',
        8: 'H',
        9: 'I',
        10: 'J',
        11: 'K',
        12: 'L',
        13: 'M',
        14: 'N',
        15: 'O',
        16: 'P',
        17: 'Q',
        18: 'R',
        19: 'S',
        20: 'T',
        21: 'U',
        22: 'V',
        23: 'W',
        24: 'X',
        25: 'Y',
        26: 'Z',
    }

    res=''
        
    prefix = 0
    flag = 0
    while(int(index/26)):
        # print(index,prefix)
        prefix+=1
        if index-26>0:
            index = index-26
        else:
            index = int((index-26)%26)
        flag = 1
    
    
    if flag==1 and index==0:
        index = 26    
        prefix-=1
    res = dic[prefix]+dic[index]
    return res


def read_from_database(folder_dir,data_file_name):
    # --- sku list
    SKU_LIST = pd.read_excel(folder_dir+data_file_name,sheet_name='MP standard list',header=5,usecols='A:AA',dtype=str)
    SKU_LIST = SKU_LIST[SKU_LIST['FG']=='FG'].reset_index(drop=True)
    SKU_LIST.fillna('',inplace=True)

    # --- vendor name
    OEM_name = pd.read_excel(folder_dir+data_file_name,sheet_name='Vendor Name',dtype=str)
    OEM_name.fillna('',inplace=True)

    # --- forecast
    forecast = pd.read_excel(folder_dir+data_file_name,sheet_name='original FCST',dtype=str,header=0)
    forecast.fillna('',inplace=True)

    index=[x for x in range(12)]
    month=list(filter(lambda x:x.find('Unnamed')==-1,forecast.columns.tolist()))[1:]
    month_dist=dict(zip(index,month))

    forecast.columns=forecast.iloc[0]
    forecast.drop(0,inplace=True)
    new_index=list(forecast.columns)
    for i in range(0,len(forecast.columns[6:])):
        forecast[forecast.columns[6+i]]=forecast[forecast.columns[6+i]].astype(int)
        if int((i)/4)>=12:
            new_index[6+i]=new_index[6+i]+'_'+str(datetime.datetime.now().year+1)+' '+month_dist.get(int((i)/4)%12)
        else:
            new_index[6+i]=new_index[6+i]+'_'+month_dist.get(int((i)/4))

    forecast=forecast.set_axis(new_index,axis='columns')

    # --- config 
    config_bp = pd.read_excel(folder_dir+data_file_name,sheet_name='Config',dtype=str,usecols='A')
    config_fc = pd.read_excel(folder_dir+data_file_name,sheet_name='Config',dtype=str,usecols='C')
    config_hcc = pd.read_excel(folder_dir+data_file_name,sheet_name='Config',dtype=str,usecols='E')

    config_bp.dropna(inplace=True)
    config_fc.dropna(inplace=True)
    config_hcc.dropna(inplace=True)

    config_bp_list = config_bp[config_bp.columns[0]].apply(lambda x:x.replace(" ",'')).tolist()
    config_fc_list = config_fc[config_fc.columns[0]].tolist()
    config_hcc_list = config_hcc[config_hcc.columns[0]].tolist()

    # --- forecast
    forecast=forecast[forecast['Branch Plant'].apply(lambda x:x.replace(" ",'')).isin(config_bp_list)]
    forecast=forecast[forecast['Model'].isin(SKU_LIST['Model'].tolist())]
    forecast=pd.merge(forecast,OEM_name[['Vendor Code','Short name (other name)']],left_on='Supplier',right_on='Vendor Code',how='left')
    del forecast['Vendor Code'],forecast['Supplier'],forecast['Supplier Name']
    forecast.rename(columns={'Short name (other name)':'Vendor'},inplace=True)
    forecast.fillna('',inplace=True)
    # print(forecast['Branch Plant'].unique())
    
    # --- BA
    customer_BA=pd.read_excel(folder_dir+data_file_name,sheet_name='Customer and BA',dtype=str,header=0)
    customer_BA.fillna('',inplace=True)

    # --- merge sheets above
    res=pd.merge(forecast,SKU_LIST[['Model','Item Status','RB Series','Category','HC Supplier','HC JDE#','3/8" medium metal ring','hub cap usage','Fabric Control','XJH Usage 210D','XJH usage 300D','XJH usage 600D']],left_on=['Model'],right_on=['Model'],how='left')
    # print(res['Branch Plant'].unique())
    res=res.drop_duplicates()
    res=pd.merge(res,customer_BA[['AB#','US BA','Customer Group','JDE FCST RP']],left_on=['Customer Num'],right_on=['AB#'],how='left')
    # print(res['Branch Plant'].unique())
    del res['AB#']
    res=res.drop_duplicates()
    res['hub cap usage']=res['hub cap usage'].apply(lambda x:0 if x=='' else x)
    res['hub cap usage']=res['hub cap usage'].apply(lambda x:int(x))
    res['XJH Usage 210D']=res['XJH Usage 210D'].apply(lambda x:0 if x=='' else x)
    res['XJH Usage 210D']=res['XJH Usage 210D'].apply(lambda x:float(x))
    res['XJH usage 300D']=res['XJH usage 300D'].apply(lambda x:0 if x=='' else x)
    res['XJH usage 300D']=res['XJH usage 300D'].apply(lambda x:float(x))
    res['XJH usage 600D']=res['XJH usage 600D'].apply(lambda x:0 if x=='' else x)
    res['XJH usage 600D']=res['XJH usage 600D'].apply(lambda x:float(x))

    res.rename(columns={'HC Supplier':'Hub Cap Supplier',
                    'HC JDE#':'Hub Cap PN#',},inplace=True)
    
    return res,config_fc_list,config_hcc_list,month_dist


def write_new_file(folder_dir,new_report_name,res,config_fc_list,config_hcc_list,month_dist):
    # -------------------excel
    writer = pd.ExcelWriter(folder_dir+new_report_name, engine='openpyxl')
    workbook = writer.book 
    font_title = Font(name="等线", size=11, bold=True)
    font_content = Font(name="等线", size=11)
    align_content = Alignment(horizontal='center',vertical='center')
    # print(res['Branch Plant'].unique())
    res.iloc[:,:80].to_excel(writer, sheet_name='FCST-This week', index=False, startrow=1)
    worksheet_fc = writer.sheets['FCST-This week']
    
    # -------CONTENT
    # variance公式设置
    for rp in range(3,len(res)+3):
        for cp in range(8,77,4):
            worksheet_fc[excel_column_formatter(cp)+str(rp)]='=IF('+excel_column_formatter(cp-1)+str(rp)+'-'+excel_column_formatter(cp-2)+str(rp)+'-'+excel_column_formatter(cp-3)+str(rp)+'>0,'+excel_column_formatter(cp-1)+str(rp)+'-'+excel_column_formatter(cp-2)+str(rp)+'-'+excel_column_formatter(cp-3)+str(rp)+',"0")'

    # 数据部分列名设置
    worksheet_fc['A1']=str(datetime.datetime.now().year)+' - Units ('+str(datetime.datetime.now().month)+'/'+str(datetime.datetime.now().day)+')'
    for cp in range(5,77,4):
        if int((cp)/4)>12:
            worksheet_fc[excel_column_formatter(cp)+'1']=str(datetime.datetime.now().year+1)+' '+month_dist.get((int(cp/4)-1)%12)
        else:
            worksheet_fc[excel_column_formatter(cp)+'1']=month_dist.get(int(cp/4)-1)
        worksheet_fc.merge_cells(excel_column_formatter(cp)+'1:'+excel_column_formatter(cp+3)+'1')
        worksheet_fc[excel_column_formatter(cp)+'2']='Open'
        worksheet_fc[excel_column_formatter(cp+1)+'2']='Shipped'
        worksheet_fc[excel_column_formatter(cp+2)+'2']='Forecast'
        worksheet_fc[excel_column_formatter(cp+3)+'2']='Variance'

    # eight new added columns 列名
    worksheet_fc[excel_column_formatter(81)+'1']='This Year'
    worksheet_fc.merge_cells(excel_column_formatter(81)+'1:'+excel_column_formatter(84)+'1')
    worksheet_fc[excel_column_formatter(81)+'2']='Total Open Order'
    worksheet_fc[excel_column_formatter(82)+'2']='Total Shipped'
    worksheet_fc[excel_column_formatter(83)+'2']='Total original forecast (Jan-Dec)'
    worksheet_fc[excel_column_formatter(84)+'2']='Valid Balance Forecast'
    worksheet_fc[excel_column_formatter(85)+'1']='Next Year'
    worksheet_fc.merge_cells(excel_column_formatter(85)+'1:'+excel_column_formatter(88)+'1')
    worksheet_fc[excel_column_formatter(85)+'2']='Total Open Order'
    worksheet_fc[excel_column_formatter(86)+'2']='Total Shipped'
    worksheet_fc[excel_column_formatter(87)+'2']='Total original forecast (Jan-Dec)'
    worksheet_fc[excel_column_formatter(88)+'2']='Valid Balance Forecast'

    res.iloc[:,80:].to_excel(writer, sheet_name='FCST-This week', index=False, startrow=1,startcol=88)

    # eight new added columns 内容
    cstart=5
    for cp in range(81,85):
        if cp==84:
            cstart = 8+(datetime.datetime.now().month-1)*4
        for rp in range(3,len(res)+2+1):
            worksheet_fc[excel_column_formatter(cp)+str(rp)]="="+(str(rp)+"+").join([excel_column_formatter(cindex) for cindex in range(cstart,53,4)])+str(rp)
        cstart+=1
    
    cstart=53
    for cp in range(85,89):
        for rp in range(3,len(res)+2+1):
            worksheet_fc[excel_column_formatter(cp)+str(rp)]="="+(str(rp)+"+").join([excel_column_formatter(cindex) for cindex in range(cstart,77,4)])+str(rp)
        cstart+=1

    # -------FORMAT
    for rp in range(1,3):
        for cell in worksheet_fc[rp]:
            cell.font = font_title
            cell.alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            if rp==2:
                cell.fill = PatternFill("solid", fgColor="D0CECE")
                
    for rp in range(3,len(res)+1+2):
        for cell in worksheet_fc[rp]:
            cell.font = font_content
            cell.alignment = align_content

    for rp in range(3,len(res)+2+1):
        for cp in range(4,18*4+8,4):
            worksheet_fc[excel_column_formatter(cp)+str(rp)].font = font_title
            if cp!=4:
                worksheet_fc[excel_column_formatter(cp)+str(rp)].fill = PatternFill("solid", fgColor="FFFF00")

    for cp in range(1,5):
        worksheet_fc[excel_column_formatter(cp)+"2"].fill = PatternFill("solid", fgColor="FFFF00")
    for cp in range(81,89):
        worksheet_fc[excel_column_formatter(cp)+"2"].fill = PatternFill("solid", fgColor="B8CCE4")
    for cp in range(89,97):
        worksheet_fc[excel_column_formatter(cp)+"2"].fill = PatternFill("solid", fgColor="DDD9C4")
    for cp in range(97,100):
        worksheet_fc[excel_column_formatter(cp)+"2"].fill = PatternFill("solid", fgColor="00B050")
        
    for cp in range(1,97):
        worksheet_fc.column_dimensions[excel_column_formatter(cp)].width=9

    # 冻结窗口:冻结对应单元格上一行和左边一列
    worksheet_fc.freeze_panes = 'E3'  # 冻结首行 两条命令后覆盖

    pd.DataFrame().to_excel(writer, sheet_name='Hub Cap Report', index=True)
    pd.DataFrame().to_excel(writer, sheet_name='Fabric Control Report', index=True)

    writer.close()

    #  -----------------python code-----------------------
    len_=len(res)
    print('length is: ',len_)

    code_file_name_hcc='Python Code For Hub Cap Control.txt'
    code_file_name_fc='Python Code For Fabric Control.txt'

    python_code_hcc="""=PY(
import pandas as pd
import datetime

forecast=xl("'FCST-This week'!A2:CK{len_}", headers=True)
forecast=forecast[forecast['Hub Cap Supplier'].isin({config_hcc_list})]
forecast=forecast.iloc[:,:80]
forecast.columns = ['Branch Plant', 'Customer Num', 'Customer Name', 'Model', 'Open_Jan',
       'Shipped_Jan', 'Forecast_Jan', 'Variance_Jan', 'Open_Feb',
       'Shipped_Feb', 'Forecast_Feb', 'Variance_Feb', 'Open_Mar',
       'Shipped_Mar', 'Forecast_Mar', 'Variance_Mar', 'Open_Apr',
       'Shipped_Apr', 'Forecast_Apr', 'Variance_Apr', 'Open_May',
       'Shipped_May', 'Forecast_May', 'Variance_May', 'Open_Jun',
       'Shipped_Jun', 'Forecast_Jun', 'Variance_Jun', 'Open_Jul',
       'Shipped_Jul', 'Forecast_Jul', 'Variance_Jul', 'Open_Aug',
       'Shipped_Aug', 'Forecast_Aug', 'Variance_Aug', 'Open_Sep',
       'Shipped_Sep', 'Forecast_Sep', 'Variance_Sep', 'Open_Oct',
       'Shipped_Oct', 'Forecast_Oct', 'Variance_Oct', 'Open_Nov',
       'Shipped_Nov', 'Forecast_Nov', 'Variance_Nov', 'Open_Dec',
       'Shipped_Dec', 'Forecast_Dec', 'Variance_Dec', 'Open_2025 Jan',
       'Shipped_2025 Jan', 'Forecast_2025 Jan', 'Variance_2025 Jan',
       'Open_2025 Feb', 'Shipped_2025 Feb', 'Forecast_2025 Feb',
       'Variance_2025 Feb', 'Open_2025 Mar', 'Shipped_2025 Mar',
       'Forecast_2025 Mar', 'Variance_2025 Mar', 'Open_2025 Apr',
       'Shipped_2025 Apr', 'Forecast_2025 Apr', 'Variance_2025 Apr',
       'Open_2025 May', 'Shipped_2025 May', 'Forecast_2025 May',
       'Variance_2025 May', 'Open_2025 Jun', 'Shipped_2025 Jun',
       'Forecast_2025 Jun', 'Variance_2025 Jun', 'Vendor', 'Item Status',
       'RB Series', 'Category']
 
cal_columns=list(filter(lambda x:x.find('Variance')!=-1,forecast.columns.tolist()))
cal_columns

res=pd.DataFrame()
for col in cal_columns:
    forecast[col].fillna(0,inplace=True)
    forecast[col]=forecast[col].astype(int)
    res['Sum of '+col.split('_')[1]]=forecast.groupby(['Vendor','RB Series'])[col].sum()

res=res.reset_index()
res
""".format(len_=len_+2,config_hcc_list=config_hcc_list)

    python_code_fc="""=PY(import pandas as pd
import datetime

forecast=xl("'FCST-This week'!A2:CO{len_}", headers=True)
forecast=forecast[forecast['Fabric Control'].isin({config_fc_list})]
forecast=forecast.iloc[:,:80]
forecast.columns = ['Branch Plant', 'Customer Num', 'Customer Name', 'Model', 'Open_Jan',
       'Shipped_Jan', 'Forecast_Jan', 'Variance_Jan', 'Open_Feb',
       'Shipped_Feb', 'Forecast_Feb', 'Variance_Feb', 'Open_Mar',
       'Shipped_Mar', 'Forecast_Mar', 'Variance_Mar', 'Open_Apr',
       'Shipped_Apr', 'Forecast_Apr', 'Variance_Apr', 'Open_May',
       'Shipped_May', 'Forecast_May', 'Variance_May', 'Open_Jun',
       'Shipped_Jun', 'Forecast_Jun', 'Variance_Jun', 'Open_Jul',
       'Shipped_Jul', 'Forecast_Jul', 'Variance_Jul', 'Open_Aug',
       'Shipped_Aug', 'Forecast_Aug', 'Variance_Aug', 'Open_Sep',
       'Shipped_Sep', 'Forecast_Sep', 'Variance_Sep', 'Open_Oct',
       'Shipped_Oct', 'Forecast_Oct', 'Variance_Oct', 'Open_Nov',
       'Shipped_Nov', 'Forecast_Nov', 'Variance_Nov', 'Open_Dec',
       'Shipped_Dec', 'Forecast_Dec', 'Variance_Dec', 'Open_2025 Jan',
       'Shipped_2025 Jan', 'Forecast_2025 Jan', 'Variance_2025 Jan',
       'Open_2025 Feb', 'Shipped_2025 Feb', 'Forecast_2025 Feb',
       'Variance_2025 Feb', 'Open_2025 Mar', 'Shipped_2025 Mar',
       'Forecast_2025 Mar', 'Variance_2025 Mar', 'Open_2025 Apr',
       'Shipped_2025 Apr', 'Forecast_2025 Apr', 'Variance_2025 Apr',
       'Open_2025 May', 'Shipped_2025 May', 'Forecast_2025 May',
       'Variance_2025 May', 'Open_2025 Jun', 'Shipped_2025 Jun',
       'Forecast_2025 Jun', 'Variance_2025 Jun', 'Vendor', 'Item Status',
       'RB Series', 'Category']
 
cal_columns=list(filter(lambda x:x.find('Variance')!=-1,forecast.columns.tolist()))
cal_columns

res=pd.DataFrame()
for col in cal_columns:
    forecast[col].fillna(0,inplace=True)
    forecast[col]=forecast[col].astype(int)
    res['Sum of '+col.split('_')[1]]=forecast.groupby(['RB Series'])[col].sum()

result = pd.DataFrame()
for col in res.columns:
    result[col] = res[col]==0
    result = result[result[col]]


res=res.reset_index()
res[~res.index.isin(result.index)]
""".format(len_=len_+2,config_fc_list=config_fc_list)

    with open(folder_dir+code_file_name_hcc, "w") as file:
        file.write(python_code_hcc)
    with open(folder_dir+code_file_name_fc, "w") as file:
        file.write(python_code_fc)


if __name__ == '__main__':
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    messagebox.showinfo("提示", "程序启动中！")

    folder_dir = os.getcwd()
    new_report_name='(DI+Dom)open_vs_forecast_units-'+str(datetime.datetime.now().month).zfill(2)+str(datetime.datetime.now().day).zfill(2)+str(datetime.datetime.now().year)+'.xlsx'
    data_file_name='FCST data base.xlsx'
    folder_dir=folder_dir+"\\"  

    res,config_fc_list,config_hcc_list,month_dist = read_from_database(folder_dir=folder_dir, data_file_name=data_file_name)
    write_new_file(folder_dir=folder_dir, new_report_name=new_report_name, res=res,config_fc_list=config_fc_list,config_hcc_list=config_hcc_list,month_dist=month_dist)    
    messagebox.showinfo("提示", "Completed.")