import pandas as pd
from tkinter import messagebox
import tkinter as tk
import os
import math
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Side, Border
from openpyxl.worksheet.page import PageMargins

in_ten_dict = {1: 'one', 2: 'two', 3: 'three', 4: 'four', 5: 'five', 6: 'six', 7: 'seven', 8: 'eight', 9: 'nine'}
in_hundred_dict = {2: "twenty", 3: "thirty", 4: "forty", 5: "fifty", 6: "sixty", 7: "seventy", 8: "eighty", 9: "ninety"}
in_twenty_dict = {10: "ten", 11: "eleven", 12: "twelve", 13: "thirteen", 14: "fourteen", 15: "fifteen", 16: "sixteen",
                  17: "seventeen", 18: "eighteen", 19: "nineteen"}


def spell_number_in_thousand(number):
    res_str = ""
    hundred_num = int(number / 100)
    if hundred_num != 0:
        hundred_num_str = in_ten_dict[hundred_num]
        res_str = res_str + hundred_num_str + ' ' + 'hundred'

    number = number % 100
    ten_num = int(number / 10)
    if ten_num == 1:
        res_str = res_str + ' ' + in_twenty_dict[number]
        return res_str
    elif ten_num != 0:
        res_str = res_str + ' ' + in_hundred_dict[ten_num]

    one_num = int(number % 10)
    if one_num != 0:
        res_str = res_str + ' ' + in_ten_dict[one_num]

    return res_str


def spell_number(number):
    res = ""

    num_float = round(math.modf(number)[0], 2) * 100
    num_int = math.modf(number)[1]

    billion_number = int(num_int / 1000000000)
    billion_str = spell_number_in_thousand(billion_number)
    if len(billion_str) != 0:
        res = res + billion_str + " billion"

    num_int = num_int % 1000000000
    million_number = int(num_int / 1000000)
    million_str = spell_number_in_thousand(million_number)
    if len(million_str) != 0:
        res = res + ' ' + million_str + " million"

    num_int = num_int % 1000000
    thousand_number = int(num_int / 1000)
    thousand_str = spell_number_in_thousand(thousand_number)
    if len(thousand_str) != 0:
        res = res + ' ' + thousand_str + " thousand"

    num_int = int(num_int % 1000)
    res = res + ' ' + spell_number_in_thousand(num_int)

    if len(res) == 0:
        res = 'zero'

    res_list = res.split(' ')
    res_list = list(filter(lambda x: len(x) != 0, res_list))
    res = " ".join(res_list)

    if not number.is_integer():
        res = res + ' AND ' + str(int(num_float)) + '/100 '
    return res.upper()


def data_extraction(file_dir, data_file_name):
    # -------------------------------read several sheetf from data file
    NON_EDI = pd.read_excel(file_dir + data_file_name,sheet_name='Non EDI',dtype=str)
    NON_EDI = NON_EDI[['Order Number','Customer PO','2nd Item Number', 'Quantity', 'First Ship Date',  'Unit Price', 'Supplier Name']]
    NON_EDI.rename(columns={'First Ship Date':'Date','Order Number':'Invoice No'},inplace=True)
    NON_EDI['Order No.']=NON_EDI['Customer PO'].apply(lambda x:x.split('-')[0])
    del NON_EDI['Customer PO']
    NON_EDI.fillna('',inplace=True)

    vendor = pd.read_excel(file_dir + data_file_name, sheet_name='Vendor', dtype=str)
    vendor.fillna('', inplace=True)
    vendor['vendor address info']=vendor['Factory Name']+";"+vendor['Address 1'].str.strip()+";"+vendor['Address 2'].str.strip()+";"\
        +vendor['Address 3'].str.strip()+";"+vendor['Address 4']

    item_master = pd.read_excel(file_dir + data_file_name, sheet_name='Item master', header=7, usecols='A:BF',
                                dtype=str)
    item_master=item_master[~item_master['Model #'].isna()]
    item_master.fillna('', inplace=True)

    item_master['sku list'] = item_master['SKU List \n(LG Software)'].apply(lambda x: x.split(','))
    item_master['sku list len'] = item_master['SKU List \n(LG Software)'].apply(lambda x: len(x.split(',')))
    single_SKU_df = item_master[item_master['sku list len'] == 1]

    mul_SKU_data = item_master[item_master['sku list len'] != 1]
    mul_SKU_df = pd.DataFrame()
    for index, rows in mul_SKU_data.iterrows():
        data = pd.DataFrame(rows)
        res = pd.concat([data.T for i in range(rows['sku list len'])]).reset_index(drop=True)
        for i in range(rows['sku list len']):
            res.loc[i, 'SKU List \n(LG Software)'] = rows['sku list'][i]
        mul_SKU_df = pd.concat([mul_SKU_df, res])
    mul_SKU_df['SKU List \n(LG Software)']=mul_SKU_df['SKU List \n(LG Software)'].apply(lambda x:x.strip())

    item_master = pd.concat([single_SKU_df, mul_SKU_df]).reset_index(drop=True)
    item_master['SKU List \n(LG Software)'] = item_master['SKU List \n(LG Software)'].apply(
        lambda x: x.replace(' ', ''))

    tcin = pd.read_excel(file_dir+data_file_name,sheet_name='TCIN# & DPCI#',dtype=str)

    # ----------------------------------merge all these data above
    df=pd.merge(NON_EDI,tcin,left_on='2nd Item Number',right_on='RF #',how='left')
    del df['RF #']

    df=pd.merge(df,vendor[['Factory ID','Factory Name','vendor address info',
                        'FOB PORT', 'Origin Country', 'JDE SUPPLIER']],left_on='Supplier Name',right_on='JDE SUPPLIER',how='left')
    df['Port of Export']=df['FOB PORT']+","+df['Origin Country']
    # df.rename(columns={'FOB':'Port of Export'},inplace=True)
    del df['FOB PORT']
    
    df = pd.merge(df,item_master,left_on=['Factory ID','2nd Item Number'],right_on=['Factory','SKU List \n(LG Software)'],how='left')

    df.fillna('',inplace=True)
    df=df.drop_duplicates(['Invoice No','2nd Item Number']).reset_index(drop=True)
    df['PO#']=df['Order No.']
    df['Age Grade (years)']=df['Age Grade (years)']+' YEARS'
    
    data=df[['Invoice No', 'Date','Payment terms', 'Order No.','Port of Export','Factory Name','vendor address info','Origin Country',
           'TCIN#', 'DPCI','PO#','Qty/\nCarton','SKU List \n(LG Software)',
             'Description','For US','Age Grade (years)','Quantity', 'Unit Price',
            'Gross Weight (kg)','Net Weight (kg)','Cubic\nMeters (per carton)','Width (L) cm', 'Depth (W) cm', 'Height (H) cm',
           'sku list', 'sku list len']]

    for i in data['Invoice No'].unique():
        temp = data[data['Invoice No'] == i]
        for j in temp['Factory Name'].unique():
            temp2 = temp[temp['Factory Name'] == j]
            model_list = temp2.drop_duplicates('SKU List \n(LG Software)')['SKU List \n(LG Software)'].to_list()
            data.loc[
                (data['Invoice No'] == i) & (data['Factory Name'] == j), 'model_list'] = 'MODLE#: ' + ', '.join(
                model_list)

    data.loc[(data['model_list'] != 'MODLE#: '), 'manufacturer'] = data['vendor address info'] + ';' + data[
        'model_list']

    return data


def data_integration(data):
    table_head_df = data[['Invoice No', 'Date','Payment terms', 'Order No.','Port of Export','Factory Name','vendor address info','Origin Country']]
    table_head_dict = table_head_df.drop_duplicates('Invoice No').set_index('Invoice No').to_dict('index')

    for i in data['Invoice No'].unique():
        temp = data[data['Invoice No'] == i]
        manufacturer_list = list(temp['manufacturer'].dropna().unique())
        table_head_dict[i].update({'manufacturer_list': manufacturer_list})

    table_content_df = data[
        ['Invoice No', 'SKU List \n(LG Software)',
        'TCIN#', 'DPCI', 'PO#','Qty/\nCarton',
        'Description','For US', 'Age Grade (years)', 
        'Quantity', 'Unit Price', 
        'Gross Weight (kg)','Net Weight (kg)', 'Cubic\nMeters (per carton)',
        'Width (L) cm','Depth (W) cm', 'Height (H) cm']]
    table_content_df = table_content_df.drop_duplicates()
    model_dict_list = []
    for i in table_content_df['Invoice No'].unique():
        temp = table_content_df[table_content_df['Invoice No'] == i]
        model_dict_list.append(
            temp[['SKU List \n(LG Software)',
                          'TCIN#', 'DPCI', 'PO#','Qty/\nCarton',
                           'Description','For US', 'Age Grade (years)', 
                          'Quantity', 'Unit Price', 
                          'Gross Weight (kg)','Net Weight (kg)', 'Cubic\nMeters (per carton)',
                          'Width (L) cm','Depth (W) cm', 'Height (H) cm']].to_dict('records'))
        # print(model_dict_list,len(model_dict_list))
    model_dict = dict(zip(table_content_df['Invoice No'].unique(), model_dict_list))
    return table_head_dict, model_dict


def write_inv_template(worksheet_inv, file_name, temp_dict, temp_content_dict):
    # 定义格式（字体、对齐方式）
    font_title = Font(name="Arail", size=9, bold=True, italic=True)
    font_content = Font(name="Arail", size=9)
    align_title = Alignment(vertical='bottom')
    align_content = Alignment(horizontal='left',vertical='bottom')
    title_row_number = 1

    # -------------------写入内容
    # ---表头
    worksheet_inv['A1'] = 'The Radio Flyer Company'
    worksheet_inv['A1'].font = Font(name="Arail", size=30, bold=True)
    worksheet_inv['A1'].alignment = Alignment(horizontal='center',vertical='center')
    worksheet_inv.merge_cells('A1:H3')
    worksheet_inv['A4'] = '6515 W GRAND AVE CHICAGO,IL 60707, USA'
    worksheet_inv['A4'].font = Font(name="Arail", size=11)
    worksheet_inv['A4'].alignment = Alignment(horizontal='center',vertical='center')
    worksheet_inv.merge_cells('A4:H5')
    for i in range(1, 6):
        worksheet_inv.row_dimensions[i].height = 15
        
    for char in range(ord('A'),ord('I')):
        side=Side(style="thin")
        worksheet_inv[chr(char)+'5'].border=Border(bottom=side)

    # ---表头左半部分
    worksheet_inv['A7'] = 'Sold To:'
    worksheet_inv['A12'] = 'Consigned to :'
    worksheet_inv['A7'].font = font_title
    worksheet_inv['A7'].alignment = align_title
    worksheet_inv['A12'].font = font_title
    worksheet_inv['A12'].alignment = align_title
    
    worksheet_inv['B7'] = 'Target Global Sourcing Ltd.'
    worksheet_inv['B8'] = '22nd Floor, One Harbourfront,'
    worksheet_inv['B9'] = '18 Tak Fung Street, Hung Hom,'
    worksheet_inv['B10'] = 'Kowloon, Hong Kong'
    worksheet_inv['B12'] = 'TARGET'
    worksheet_inv['B13'] = '7000 Target Parkway North'
    worksheet_inv['B14'] = 'Brooklyn Park, Minnesota 55445'
    worksheet_inv['B15'] = 'Tel: 763-405-0296 Fax: 612-304-3113'
    for i in range(7,16):
        worksheet_inv['B' + str(i)].font = font_content
        worksheet_inv['B' + str(i)].alignment = align_content

    # ---表头右半部分
    worksheet_inv['D7'] = 'Invoice No:'
    worksheet_inv['D8'] = 'Date:'
    worksheet_inv['D9'] = 'Payment Terms:'
    worksheet_inv['D10'] = 'Order No.'
    worksheet_inv['D11'] = 'Port of Export:'
    worksheet_inv['D13'] = 'FCA ---'
    for i in range(7,14):
        worksheet_inv['D' + str(i)].font = font_title
        worksheet_inv['D' + str(i)].alignment = align_title
        
    worksheet_inv['E7'] = file_name
    date_str = temp_dict['Date'][:-9]
    worksheet_inv['E8'] = str(int(date_str.split('-')[1]))+'/'+str(int(date_str.split('-')[2]))+'/'+str(int(date_str.split('-')[0]))
    worksheet_inv['E9'] = temp_dict['Payment terms']
    worksheet_inv['E10'] = temp_dict['Order No.']
    worksheet_inv['E11'] = temp_dict['Port of Export']
    for i in range(7,12):
        worksheet_inv['E' + str(i)].font = font_content
        worksheet_inv['E' + str(i)].alignment = align_content
    
    list_pointer=0
    writer_pointer=13
    while(list_pointer<len(temp_dict['manufacturer_list'])):
        info=temp_dict['manufacturer_list'][list_pointer]
        if info!='':
            worksheet_inv['E' + str(writer_pointer)]=info
            worksheet_inv['E' + str(writer_pointer)].font = Font(name="Arail", size=8)
            worksheet_inv.merge_cells("E"+str(writer_pointer)+':H'+str(writer_pointer+2))
            worksheet_inv['E' + str(writer_pointer)].alignment = Alignment(horizontal='left',vertical='top',wrapText=True)

            # 待验证
            writer_pointer=writer_pointer+3
        list_pointer=list_pointer+1

    # writer_pointer=writer_pointer+1
    worksheet_inv['A' + str(writer_pointer)] = 'Shipped Via:'
    worksheet_inv['A' + str(writer_pointer)].font = font_title
    worksheet_inv['A' + str(writer_pointer)].alignment = align_title
    
    worksheet_inv['B' + str(writer_pointer)] = 'A vessel'
    worksheet_inv['B' + str(writer_pointer)].font = font_content
    worksheet_inv['B' + str(writer_pointer)].alignment = align_content

    worksheet_inv['D' + str(writer_pointer)] = 'Country of Origin:'
    worksheet_inv['D' + str(writer_pointer)].font = font_title
    worksheet_inv['D' + str(writer_pointer)].alignment = align_title
    # worksheet_inv['B' + str(writer_pointer)] = 'China'
    worksheet_inv['E' + str(writer_pointer)] = temp_dict['Origin Country']
    worksheet_inv['E' + str(writer_pointer)].font = font_content
    worksheet_inv['E' + str(writer_pointer)].alignment = align_content

    writer_pointer=writer_pointer+1
    worksheet_inv['D' + str(writer_pointer)] = 'Final Destination:'
    worksheet_inv['D' + str(writer_pointer)].font = font_title
    worksheet_inv['D' + str(writer_pointer)].alignment = align_title

    worksheet_inv['E' + str(writer_pointer)] = 'U.S.A.'
    worksheet_inv['E' + str(writer_pointer)].font = font_content
    worksheet_inv['E' + str(writer_pointer)].alignment = align_content

    
    # ---表格部分！！！
    font_head_title=Font(name="Arail", size=9, bold=True)
    font_important_content = Font(name="Arail", size=9, bold=True)
    font_trivial_content = Font(name="Arail", size=9)
    
    # 表格表头
    writer_pointer=writer_pointer+1
    worksheet_inv['A'+str(writer_pointer)] = 'Marks & Nos.'
    worksheet_inv.merge_cells("A"+str(writer_pointer)+':B'+str(writer_pointer))
    worksheet_inv['C'+str(writer_pointer)] = 'Description'
    worksheet_inv.merge_cells("C"+str(writer_pointer)+':D'+str(writer_pointer))
    worksheet_inv['E'+str(writer_pointer)] = 'Quantity'
    worksheet_inv['F'+str(writer_pointer)] = 'Cartons'
    worksheet_inv['G'+str(writer_pointer)] = 'Unit Price'
    worksheet_inv['H'+str(writer_pointer)] = 'Amount'
    for cell in worksheet_inv[writer_pointer]:
        cell.font = font_head_title
        cell.alignment = Alignment(horizontal='center',vertical='center')
    worksheet_inv.row_dimensions[writer_pointer].height = 12

    # 设置线条的样式和颜色
    side = Side(style="medium")
    # 设置单元格的边框线条
    worksheet_inv['A'+str(writer_pointer)].border = Border(bottom=side,top=side)
    worksheet_inv['B'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_inv['C'+str(writer_pointer)].border = Border(bottom=side,top=side)
    worksheet_inv['D'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_inv['E'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_inv['F'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_inv['G'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_inv['H'+str(writer_pointer)].border = Border(bottom=side,top=side)
    
    # 表格内容
    model_list = temp_content_dict
    model_start_pointer = writer_pointer+2
    title_row_number = writer_pointer + 1
    
    total_amount = 0
    item_count = 0
    for info in model_list:
        item_count += 1
        if item_count == 8:
            worksheet_inv.print_title_rows = '1:' + str(title_row_number)
            writer_pointer = 66 
        else:
            writer_pointer = writer_pointer + 2
                
        i=writer_pointer
        
        print(info)
        worksheet_inv['A'+str(writer_pointer)] = 'DPCI:'
        worksheet_inv['B'+str(writer_pointer)] = info['DPCI']
        worksheet_inv['C'+str(writer_pointer)] = 'Item No.:'
        worksheet_inv['D'+str(writer_pointer)] = '#'+info['SKU List \n(LG Software)']+' - '+info['Description']
        worksheet_inv['E'+str(writer_pointer)] = int(info['Quantity'])
        worksheet_inv['F'+str(writer_pointer)] = int(info['Quantity'])/int(info['Qty/\nCarton'])
        worksheet_inv['G'+str(writer_pointer)] = float(info['Unit Price'])
        worksheet_inv['H'+str(writer_pointer)] = '=E'+str(writer_pointer)+'*G'+str(writer_pointer)
        worksheet_inv['G'+str(writer_pointer)].number_format='"$"#,##0.00_-'
        worksheet_inv['H'+str(writer_pointer)].number_format='"$"#,##0.00_-'
        total_amount = total_amount+int(info['Quantity'])*float(info['Unit Price'])
        
        writer_pointer=writer_pointer+1
        worksheet_inv['A'+str(writer_pointer)] = 'PO#:'
        worksheet_inv['B'+str(writer_pointer)] = info['PO#']
        worksheet_inv['C'+str(writer_pointer)] = 'TCIN#:'
        worksheet_inv['D'+str(writer_pointer)] = info['TCIN#']
        writer_pointer=writer_pointer+1
        worksheet_inv['A'+str(writer_pointer)] = 'VCP/SSP:'
        worksheet_inv['B'+str(writer_pointer)] = info['Qty/\nCarton']+'/'+info['Qty/\nCarton']
        worksheet_inv['C'+str(writer_pointer)] = 'DPCI:'
        worksheet_inv['D'+str(writer_pointer)] = info['DPCI']
        writer_pointer=writer_pointer+1
        worksheet_inv['A'+str(writer_pointer)] = 'ITEM:'
        worksheet_inv['B'+str(writer_pointer)] = info['SKU List \n(LG Software)']
        worksheet_inv['C'+str(writer_pointer)] = 'HTS#:'
        worksheet_inv['D'+str(writer_pointer)] = info['For US']
        writer_pointer=writer_pointer+1
        worksheet_inv['C'+str(writer_pointer)] = 'AGE RANGE:'
        worksheet_inv['D'+str(writer_pointer)] = info['Age Grade (years)']

        for char in range(ord('A'),ord('I')):
            for j in range(i,writer_pointer+1):
                worksheet_inv[chr(char)+str(j)].font=font_trivial_content
                if chr(char)=='E' or chr(char)=='F' or chr(char)=='G' or chr(char)=='H':
                    worksheet_inv[chr(char)+str(j)].alignment = Alignment(horizontal='center',vertical='center')
                else:
                    worksheet_inv[chr(char)+str(j)].alignment = Alignment(horizontal='left',vertical='center')
    
    model_end_pointer = writer_pointer
    writer_pointer=writer_pointer+2
    
    #取出表中的各行,第二行到倒数第二行
    for char in range(ord('A'),ord('H')):
        char=chr(char)
        if char=='A' or char=='C':
            continue
        for i in range(model_start_pointer-1,writer_pointer):
            worksheet_inv[char+str(i)].border=Border(right=side)
    
    worksheet_inv['C'+str(writer_pointer)] = 'TOTAL'
    worksheet_inv['E'+str(writer_pointer)] = '=SUM(E'+str(model_start_pointer)+':E'+str(model_end_pointer)+')'
    worksheet_inv['F'+str(writer_pointer)] = '=SUM(F'+str(model_start_pointer)+':F'+str(model_end_pointer)+')'
    worksheet_inv['H'+str(writer_pointer)] = '=ROUND(SUM(H'+str(model_start_pointer)+':H'+str(model_end_pointer)+'),2)'
    
    for char in range(ord('A'),ord('I')):
        worksheet_inv[chr(char)+str(writer_pointer)].border=Border(top=side)
    worksheet_inv['H'+str(writer_pointer)].number_format='"$"#,##0.00_-'
    
    for cell in worksheet_inv[writer_pointer]:
        cell.font = font_important_content
        cell.alignment = Alignment(horizontal='center',vertical='center')
    
    for i in range(model_start_pointer-1, writer_pointer+1):
        worksheet_inv.row_dimensions[i].height = 12
        
    
    # ---表格尾部
    
    worksheet_inv['C'+str(writer_pointer+2)] = 'SAY: '+spell_number(total_amount)+' U.S. DOLLARS ONLY.'
    worksheet_inv['C'+str(writer_pointer+2)].font=Font(name="Arail", size=9, bold=True)
    worksheet_inv['C'+str(writer_pointer+4)] = "SHIPPER'S DECLARATION CONCERNING WOOD PACKING MATERIALS:"
    worksheet_inv['C'+str(writer_pointer+5)] = 'NO WOOD PACKING MATERIAL IS USED IN THE SHIPMENT'
    worksheet_inv['C'+str(writer_pointer+4)].font=Font(name="Arail", size=9)
    worksheet_inv['C'+str(writer_pointer+5)].font=Font(name="Arail", size=9)
    worksheet_inv['C'+str(writer_pointer+7)] = 'We hereby certify that all goods have been marked in accordance with U.S. laws, rules and regulations, including CBP (Department of Homeland Security Bureau of Customs and Border Protection) laws pertaining to Country of Origin markings. '
    worksheet_inv.merge_cells("C"+str(writer_pointer+7)+':H'+str(writer_pointer+8))
    worksheet_inv['C'+str(writer_pointer+7)].font=Font(name="Arail", size=8, bold=True)
    for i in range(writer_pointer+2,writer_pointer+8):
        worksheet_inv['C'+str(i)].alignment = Alignment(horizontal='left',vertical='center')
    worksheet_inv['C'+str(writer_pointer+7)].alignment=Alignment(wrapText=True)
    
    worksheet_inv['E'+str(writer_pointer+10)] = 'The Radio Flyer Company'
    worksheet_inv['E'+str(writer_pointer+10)].font=Font(name="Arail", size=9, bold=True, italic=True)

    # 设置单元格的边框线条
    border = Border(bottom=side)
    worksheet_inv['E'+str(writer_pointer+15)].border = border
    worksheet_inv['F'+str(writer_pointer+15)].border = border
    worksheet_inv['G'+str(writer_pointer+15)].border = border
    worksheet_inv['H'+str(writer_pointer+15)].border = border
    worksheet_inv['E'+str(writer_pointer+16)] = 'Rainbow Lin'
    worksheet_inv['E'+str(writer_pointer+17)] = 'Logistics Planner'
    worksheet_inv['E'+str(writer_pointer+16)].font=Font(name="Arail", size=9, bold=True, italic=True)
    worksheet_inv['E'+str(writer_pointer+17)].font=Font(name="Arail", size=9, bold=True, italic=True)

    manufacturer_list=temp_dict['manufacturer_list']
    for info in manufacturer_list:
        writer_pointer=writer_pointer+2
        worksheet_inv['A'+str(writer_pointer)] = 'Manufacturer:'
        worksheet_inv['A'+str(writer_pointer)].font=font_important_content
        
        info_list=info.split(';')
        for i in info_list:
            if i!='':
                writer_pointer=writer_pointer+1
                if i.find('MODLE#')!=-1:
                    if len(i)>40:
                        pos=len(i)
                        while pos>40:
                            pos=i.rfind(', ',0,pos)
                        worksheet_inv['A'+str(writer_pointer)] = i[:pos+2]
                        worksheet_inv['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
                        
                        writer_pointer=writer_pointer+1
                        worksheet_inv['A'+str(writer_pointer)] = i[pos+2:]                    
                        worksheet_inv['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
                    else:
                        worksheet_inv['A'+str(writer_pointer)] = i
                        worksheet_inv['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
                else:
                    worksheet_inv['A'+str(writer_pointer)] = i
                    worksheet_inv['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
    
    
    writer_pointer=writer_pointer+2
    worksheet_inv['A'+str(writer_pointer)] = "Seller's name & address:"
    worksheet_inv['A'+str(writer_pointer)].font=font_important_content
    worksheet_inv['A'+str(writer_pointer+1)] = 'The Radio Flyer Company'+'\n'+'6515 W Grand Ave., Chicago IL 60707, USA'
    worksheet_inv['A'+str(writer_pointer+1)].font=Font(name="Arail", size=9)
    worksheet_inv['A'+str(writer_pointer+1)].alignment=Alignment(wrapText=True)
    worksheet_inv.merge_cells("A"+str(writer_pointer+1)+':B'+str(writer_pointer+3))

    # ---格式调整
    
    # 调整列宽
    worksheet_inv.column_dimensions['A'].width=15
    worksheet_inv.column_dimensions['B'].width=14.8
    worksheet_inv.column_dimensions['C'].width=19
    worksheet_inv.column_dimensions['D'].width=34.5
    worksheet_inv.column_dimensions['E'].width=11.5
    worksheet_inv.column_dimensions['F'].width=11.5
    worksheet_inv.column_dimensions['G'].width=11.5
    worksheet_inv.column_dimensions['H'].width=13
    
    
    # 调整页面
    worksheet_inv.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet_inv.page_setup.fitToWidth = 1
    worksheet_inv.page_setup.fitToHeight = 0

    # 设置页眉
    worksheet_inv.oddHeader.right.text = "ORIGINAL\nCOMMERCIAL INVOICE"  # 文本
    worksheet_inv.oddHeader.right.size = 10  # 字号
    worksheet_inv.oddHeader.right.font = "Arial,Bold"  # 字体
    # worksheet_inv.oddFooter.center.text = "Page &[Page] of &[Pages]"  # 文本
    # worksheet_inv.oddFooter.center.size = 11  # 字号
    # worksheet_inv.oddFooter.center.font = "宋体"  # 字体
    
    # 设置页边距(单位:英寸)
    worksheet_inv.page_margins = PageMargins(top=1.1, bottom=1.1, left=0.25,right=0.25) 

    return worksheet_inv


def write_pl_template(worksheet_pl, file_name, temp_dict, temp_content_dict):
    # 定义格式（字体、对齐方式）
    font_title = Font(name="Arail", size=9, bold=True, italic=True)
    font_content = Font(name="Arail", size=9)
    align_title = Alignment(vertical='bottom')
    align_content = Alignment(horizontal='left',vertical='bottom')
    title_row_number = 1

    # -------------------写入内容
    # ---表头
    worksheet_pl['A1'] = 'The Radio Flyer Company'
    worksheet_pl['A1'].font = Font(name="Arail", size=30, bold=True)
    worksheet_pl['A1'].alignment = Alignment(horizontal='center',vertical='center')
    worksheet_pl.merge_cells('A1:I4')
    worksheet_pl['A5'] = '6515 W GRAND AVE CHICAGO,IL 60707, USA'
    worksheet_pl['A5'].font = Font(name="Arail", size=12)
    worksheet_pl['A5'].alignment = Alignment(horizontal='center',vertical='center')
    worksheet_pl.merge_cells('A5:I5')
    for i in range(1, 5):
        worksheet_pl.row_dimensions[i].height = 18
    for char in range(ord('A'),ord('J')):
        side=Side(style="thin")
        worksheet_pl[chr(char)+'5'].border=Border(bottom=side)

    # ---表头左半部分
    worksheet_pl['A8'] = 'Sold To:'
    worksheet_pl['A13'] = 'Consigned to :'
    worksheet_pl['A8'].font = font_title
    worksheet_pl['A8'].alignment = align_title
    worksheet_pl['A13'].font = font_title
    worksheet_pl['A13'].alignment = align_title
    
    worksheet_pl['B8'] = 'Target Global Sourcing Ltd.'
    worksheet_pl['B9'] = '22nd Floor, One Harbourfront,'
    worksheet_pl['B10'] = '18 Tak Fung Street, Hung Hom,'
    worksheet_pl['B11'] = 'Kowloon, Hong Kong'
    worksheet_pl['B13'] = 'TARGET'
    worksheet_pl['B14'] = '7000 Target Parkway North'
    worksheet_pl['B15'] = 'Brooklyn Park, Minnesota 55445'
    worksheet_pl['B16'] = 'Tel: 763-405-0296 Fax: 612-304-3113'
    for i in range(8,17):
        worksheet_pl['B' + str(i)].font = font_content
        worksheet_pl['B' + str(i)].alignment = align_content

    # ---表头右半部分
    worksheet_pl['F8'] = 'Invoice No:'
    worksheet_pl['F9'] = 'Date:'
    worksheet_pl['F10'] = 'Payment Terms:'
    worksheet_pl['F11'] = 'Order No.'
    worksheet_pl['F12'] = 'Port of Export:'
    worksheet_pl['F14'] = 'FCA ---'
    for i in range(8,15):
        worksheet_pl['F' + str(i)].font = font_title
        worksheet_pl['F' + str(i)].alignment = align_title
        
    worksheet_pl['G8'] = file_name
    date_str = temp_dict['Date'][:-9]
    worksheet_pl['G9'] = str(int(date_str.split('-')[1]))+'/'+str(int(date_str.split('-')[2]))+'/'+str(int(date_str.split('-')[0]))
    worksheet_pl['G10'] = temp_dict['Payment terms']
    worksheet_pl['G11'] = temp_dict['Order No.']
    worksheet_pl['G12'] = temp_dict['Port of Export']
    for i in range(8,13):
        worksheet_pl['G' + str(i)].font = font_content
        worksheet_pl['G' + str(i)].alignment = align_content
    
    list_pointer=0
    writer_pointer=14
    while(list_pointer<len(temp_dict['manufacturer_list'])):
        info=temp_dict['manufacturer_list'][list_pointer]
        if info!='':
            worksheet_pl['G' + str(writer_pointer)]=info
            worksheet_pl['G' + str(writer_pointer)].font = Font(name="Arail", size=8)
            worksheet_pl.merge_cells("G"+str(writer_pointer)+':I'+str(writer_pointer+2))
            worksheet_pl['G' + str(writer_pointer)].alignment = Alignment(horizontal='left',vertical='top',wrapText=True)

            # 待验证
            writer_pointer=writer_pointer+3
        list_pointer=list_pointer+1

    # writer_pointer=writer_pointer+1
    worksheet_pl['A' + str(writer_pointer)] = 'Shipped Via:'
    worksheet_pl['A' + str(writer_pointer)].font = font_title
    worksheet_pl['A' + str(writer_pointer)].alignment = align_title
    
    worksheet_pl['B' + str(writer_pointer)] = 'A vessel'
    worksheet_pl['B' + str(writer_pointer)].font = font_content
    worksheet_pl['B' + str(writer_pointer)].alignment = align_content
    
    worksheet_pl['F' + str(writer_pointer)] = 'Country of Origin:'
    worksheet_pl['F' + str(writer_pointer)].font = font_title
    worksheet_pl['F' + str(writer_pointer)].alignment = align_title
    worksheet_pl['G' + str(writer_pointer)] = temp_dict['Origin Country']
    worksheet_pl['G' + str(writer_pointer)].font = font_content
    worksheet_pl['G' + str(writer_pointer)].alignment = align_content

    writer_pointer=writer_pointer+1
    
    worksheet_pl['F' + str(writer_pointer)] = 'Final Destination:'
    worksheet_pl['F' + str(writer_pointer)].font = font_title
    worksheet_pl['F' + str(writer_pointer)].alignment = align_title
    worksheet_pl['G' + str(writer_pointer)] = 'U.S.A.'
    worksheet_pl['G' + str(writer_pointer)].font = font_content
    worksheet_pl['G' + str(writer_pointer)].alignment = align_content
    
    # ---表格部分！！！
    font_head_title=Font(name="Arail", size=9, bold=True)
    font_important_content = Font(name="Arail", size=9, bold=True)
    font_trivial_content = Font(name="Arail", size=9)
    
    # 表格表头
    writer_pointer=writer_pointer+1
    worksheet_pl['A'+str(writer_pointer)] = 'Marks & Nos.'
    worksheet_pl.merge_cells("A"+str(writer_pointer)+':B'+str(writer_pointer))
    worksheet_pl['C'+str(writer_pointer)] = 'Description'
    worksheet_pl.merge_cells("C"+str(writer_pointer)+':F'+str(writer_pointer))
    worksheet_pl['G'+str(writer_pointer)] = 'Quantity'
    worksheet_pl['H'+str(writer_pointer)] = 'Cartons'
    worksheet_pl['I'+str(writer_pointer)] = 'Weight(KG)'
    for cell in worksheet_pl[writer_pointer]:
        cell.font = font_head_title
        cell.alignment = Alignment(horizontal='center',vertical='center')
    worksheet_pl.row_dimensions[writer_pointer].height = 12

    # 设置线条的样式和颜色
    side = Side(style="medium")
    # 设置单元格的边框线条
    worksheet_pl['A'+str(writer_pointer)].border = Border(bottom=side,top=side)
    worksheet_pl['B'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_pl['C'+str(writer_pointer)].border = Border(bottom=side,top=side)
    worksheet_pl['D'+str(writer_pointer)].border = Border(bottom=side,top=side)
    worksheet_pl['E'+str(writer_pointer)].border = Border(bottom=side,top=side)
    worksheet_pl['F'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_pl['G'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_pl['H'+str(writer_pointer)].border = Border(bottom=side,top=side,right=side)
    worksheet_pl['I'+str(writer_pointer)].border = Border(bottom=side,top=side)

    # 表格内容
    model_list=temp_content_dict
    model_start_pointer = writer_pointer+2
    title_row_number = writer_pointer + 1
    model_index_list=[]
    
    carton_total=0
    item_count=0
    for info in model_list:
        item_count += 1
        if item_count == 7:
            worksheet_pl.print_title_rows = '1:' + str(title_row_number)
            writer_pointer = 96 
        else:
            writer_pointer=writer_pointer + 2

        i=writer_pointer
        
        worksheet_pl['A'+str(writer_pointer)] = 'DPCI:'
        worksheet_pl['B'+str(writer_pointer)] = info['DPCI']
        worksheet_pl['C'+str(writer_pointer)] = 'Item No.:'
        worksheet_pl['D'+str(writer_pointer)] = '#'+info['SKU List \n(LG Software)']
        worksheet_pl['E'+str(writer_pointer)] = '-'
        worksheet_pl['F'+str(writer_pointer)] = info['Description']
        worksheet_pl['G'+str(writer_pointer)] = int(info['Quantity'])
        worksheet_pl['H'+str(writer_pointer)] = int(info['Quantity'])/int(info['Qty/\nCarton'])
        worksheet_pl['I'+str(writer_pointer)] = '=F'+str(writer_pointer+6)
        worksheet_pl['I'+str(writer_pointer)].number_format='0.00'
        carton_total = carton_total+ int(info['Quantity'])/int(info['Qty/\nCarton'])
        model_index_list.append(writer_pointer)
        
        writer_pointer=writer_pointer+1
        worksheet_pl['A'+str(writer_pointer)] = 'PO#:'
        worksheet_pl['B'+str(writer_pointer)] = info['PO#']
        worksheet_pl['C'+str(writer_pointer)] = 'TCIN#:'
        worksheet_pl['D'+str(writer_pointer)] = info['TCIN#']
        writer_pointer=writer_pointer+1
        worksheet_pl['A'+str(writer_pointer)] = 'VCP/SSP:'
        worksheet_pl['B'+str(writer_pointer)] = info['Qty/\nCarton']+'/'+info['Qty/\nCarton']
        worksheet_pl['C'+str(writer_pointer)] = 'DPCI:'
        worksheet_pl['D'+str(writer_pointer)] = info['DPCI']
        writer_pointer=writer_pointer+1
        worksheet_pl['A'+str(writer_pointer)] = 'ITEM:'
        worksheet_pl['B'+str(writer_pointer)] = info['SKU List \n(LG Software)']
        worksheet_pl['C'+str(writer_pointer)] = 'HTS#:'
        worksheet_pl['D'+str(writer_pointer)] = info['For US']
        writer_pointer=writer_pointer+1
        # worksheet_pl['A'+str(writer_pointer)] = 'ITEM:'
        # worksheet_pl['B'+str(writer_pointer)] = info['SKU List \n(LG Software)']
        worksheet_pl['C'+str(writer_pointer)] = 'AGE RANGE:'
        worksheet_pl['D'+str(writer_pointer)] = info['Age Grade (years)']

        writer_pointer=writer_pointer+1
        writer_pointer=writer_pointer+1
        worksheet_pl['C'+str(writer_pointer)] = 'Gross Weight (KGS):'
        worksheet_pl['D'+str(writer_pointer)] = float(info['Gross Weight (kg)'])
        worksheet_pl['E'+str(writer_pointer)] = '/'
        # worksheet_pl['F'+str(writer_pointer)] = float(info['Gross Weight (kg)']*(int(info['Quantity'])/int(info['Qty/\nCarton'])))
        worksheet_pl['F'+str(writer_pointer)] = '=D'+str(writer_pointer)+'*H'+str(writer_pointer-6)
        worksheet_pl['D'+str(writer_pointer)].number_format='0.00'
        worksheet_pl['F'+str(writer_pointer)].number_format='0.00'
        
        writer_pointer=writer_pointer+1
        worksheet_pl['C'+str(writer_pointer)] = 'Net Weight  (KGS):'
        worksheet_pl['D'+str(writer_pointer)] = float(info['Net Weight (kg)'])
        worksheet_pl['E'+str(writer_pointer)] = '/'
        # worksheet_pl['F'+str(writer_pointer)] = float(info['Net Weight (kg)']*(int(info['Quantity'])/int(info['Qty/\nCarton'])))
        worksheet_pl['F'+str(writer_pointer)] = '=D'+str(writer_pointer)+'*H'+str(writer_pointer-7)
        worksheet_pl['D'+str(writer_pointer)].number_format='0.00'
        worksheet_pl['F'+str(writer_pointer)].number_format='0.00'
        
        writer_pointer=writer_pointer+1
        worksheet_pl['C'+str(writer_pointer)] = 'Cubic Meter (CBM):'
        worksheet_pl['D'+str(writer_pointer)] = float(info['Cubic\nMeters (per carton)'])
        worksheet_pl['E'+str(writer_pointer)] = '/'
        worksheet_pl['F'+str(writer_pointer)] = '=ROUND(D'+str(writer_pointer)+'*H'+str(writer_pointer-8)+',3)'
        worksheet_pl['D'+str(writer_pointer)].number_format='0.0000'
        worksheet_pl['F'+str(writer_pointer)].number_format='0.000'
        
        writer_pointer=writer_pointer+1
        worksheet_pl['C'+str(writer_pointer)] = 'Dimension(CM):'
        worksheet_pl['D'+str(writer_pointer)] = info['Width (L) cm']+'*'+info['Depth (W) cm']+'*'+info['Height (H) cm']
        

        for char in range(ord('A'),ord('J')):
            for j in range(i,writer_pointer+1):
                worksheet_pl[chr(char)+str(j)].font=font_trivial_content
                if chr(char)=='D'or chr(char)=='E' or chr(char)=='F' or chr(char)=='G' or chr(char)=='H' or chr(char)=='I':
                    worksheet_pl[chr(char)+str(j)].alignment = Alignment(horizontal='center',vertical='center')
                else:
                    worksheet_pl[chr(char)+str(j)].alignment = Alignment(horizontal='left',vertical='center')
    
    model_end_pointer = writer_pointer
    writer_pointer=writer_pointer+2

    worksheet_pl['C'+str(writer_pointer)] = 'Total Gross Weight (KGS):'
    worksheet_pl['D'+str(writer_pointer)] = '=SUM(F'+',F'.join(str(index+6) for index in model_index_list)+')'
    worksheet_pl['C'+str(writer_pointer)].font=font_trivial_content
    worksheet_pl['D'+str(writer_pointer)].font=font_important_content
    worksheet_pl['D'+str(writer_pointer)].number_format='0.00'
    writer_pointer=writer_pointer+1
    worksheet_pl['C'+str(writer_pointer)] = 'Total Net Weight  (KGS):'
    worksheet_pl['D'+str(writer_pointer)] = '=SUM(F'+',F'.join(str(index+7) for index in model_index_list)+')'
    worksheet_pl['C'+str(writer_pointer)].font=font_trivial_content
    worksheet_pl['D'+str(writer_pointer)].font=font_important_content
    worksheet_pl['D'+str(writer_pointer)].number_format='0.00'
    writer_pointer=writer_pointer+1
    worksheet_pl['C'+str(writer_pointer)] = 'Total Cubic Meter (CBM):'
    worksheet_pl['D'+str(writer_pointer)] = '=SUM(F'+',F'.join(str(index+8) for index in model_index_list)+')'
    worksheet_pl['C'+str(writer_pointer)].font=font_trivial_content
    worksheet_pl['D'+str(writer_pointer)].font=font_important_content
    worksheet_pl['D'+str(writer_pointer)].number_format='0.000'
    writer_pointer=writer_pointer+1
    
    #取出表中的各行,第二行到倒数第二行
    for char in range(ord('A'),ord('I')):
        char=chr(char)
        if char=='A' or char=='C' or char=='D' or char=='E' :
            continue
        for i in range(model_start_pointer-1,writer_pointer+1):
            worksheet_pl[char+str(i)].border=Border(right=side)

    writer_pointer=writer_pointer+1
    worksheet_pl['C'+str(writer_pointer)] = 'TOTAL'
    worksheet_pl['G'+str(writer_pointer)] = '=SUM(G'+',G'.join(str(index) for index in model_index_list)+')'
    worksheet_pl['H'+str(writer_pointer)] = '=SUM(H'+',H'.join(str(index) for index in model_index_list)+')'
    worksheet_pl['I'+str(writer_pointer)] = '=ROUND(SUM(I'+',I'.join(str(index) for index in model_index_list)+'),2)'
    
    for char in range(ord('A'),ord('J')):
        worksheet_pl[chr(char)+str(writer_pointer)].border=Border(top=side)
    worksheet_pl['I'+str(writer_pointer)].number_format='#,##0.00_-'
    
    for cell in worksheet_pl[writer_pointer]:
        cell.font = font_important_content
        cell.alignment = Alignment(horizontal='center',vertical='center')
    
    for i in range(model_start_pointer-1, writer_pointer+1):
        worksheet_pl.row_dimensions[i].height = 12

    # ---表格尾部
    worksheet_pl['F'+str(writer_pointer+2)] = 'SAY: '+spell_number(carton_total)+' CARTONS ONLY.'
    worksheet_pl['F'+str(writer_pointer+2)].font=Font(name="Arail", size=9, bold=True)
    worksheet_pl['F'+str(writer_pointer+4)] = "SHIPPER'S DECLARATION CONCERNING WOOD PACKING MATERIALS:"
    worksheet_pl['F'+str(writer_pointer+5)] = 'NO WOOD PACKING MATERIAL IS USED IN THE SHIPMENT'
    worksheet_pl['F'+str(writer_pointer+4)].font=Font(name="Arail", size=9)
    worksheet_pl['F'+str(writer_pointer+5)].font=Font(name="Arail", size=9)
    worksheet_pl['F'+str(writer_pointer+7)] = 'We hereby certify that all goods have been marked in accordance with U.S. laws, rules and regulations, including CBP (Department of Homeland Security Bureau of Customs and Border Protection) laws pertaining to Country of Origin markings. '
    worksheet_pl.merge_cells("F"+str(writer_pointer+7)+':I'+str(writer_pointer+8))
    worksheet_pl['F'+str(writer_pointer+7)].font=Font(name="Arail", size=8, bold=True)
    for i in range(writer_pointer+7, writer_pointer+8):
        worksheet_pl.row_dimensions[i].height = 20
        
    for i in range(writer_pointer+2,writer_pointer+9):
        worksheet_pl['F'+str(i)].alignment = Alignment(horizontal='left',vertical='center')
    worksheet_pl['F'+str(writer_pointer+7)].alignment=Alignment(wrapText=True)
    
    worksheet_pl['F'+str(writer_pointer+10)] = 'The Radio Flyer Company'
    worksheet_pl['F'+str(writer_pointer+10)].font=Font(name="Arail", size=9, bold=True, italic=True)

    # 设置单元格的边框线条
    border = Border(bottom=side)
    worksheet_pl['F'+str(writer_pointer+14)].border = border
    worksheet_pl['G'+str(writer_pointer+14)].border = border
    worksheet_pl['H'+str(writer_pointer+14)].border = border
    worksheet_pl['F'+str(writer_pointer+15)] = 'Rainbow Lin'
    worksheet_pl['F'+str(writer_pointer+16)] = 'Logistics Planner'
    worksheet_pl['F'+str(writer_pointer+15)].font=Font(name="Arail", size=9, bold=True, italic=True)
    worksheet_pl['F'+str(writer_pointer+16)].font=Font(name="Arail", size=9, bold=True, italic=True)

    writer_pointer=writer_pointer+2
    worksheet_pl['A'+str(writer_pointer)] = "Goods have been marked "
    worksheet_pl['A'+str(writer_pointer+1)] = 'in accordance with U.S.'
    worksheet_pl['A'+str(writer_pointer+2)] = 'Customs regulations.'
    for i in range(writer_pointer,writer_pointer+3):
        worksheet_pl['A'+str(i)].font = Font(name="Arail", size=9)
        worksheet_pl['A'+str(i)].alignment = Alignment(horizontal='left',vertical='center')

    writer_pointer=writer_pointer+4
    manufacturer_list=temp_dict['manufacturer_list']
    for info in manufacturer_list:
        writer_pointer=writer_pointer+2
        worksheet_pl['A'+str(writer_pointer)] = 'Manufacturer:'
        worksheet_pl['A'+str(writer_pointer)].font=font_important_content
        
        info_list=info.split(';')
        for i in info_list:
            if i!='':
                writer_pointer=writer_pointer+1
                if i.find('MODLE#')!=-1:
                    if len(i)>40:
                        pos=len(i)
                        while pos>40:
                            pos=i.rfind(', ',0,pos)
                        worksheet_pl['A'+str(writer_pointer)] = i[:pos+2]
                        worksheet_pl['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
                        
                        writer_pointer=writer_pointer+1
                        worksheet_pl['A'+str(writer_pointer)] = i[pos+2:]
                        worksheet_pl['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
                    else:
                        worksheet_pl['A'+str(writer_pointer)] = i
                        worksheet_pl['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
                else:
                    worksheet_pl['A'+str(writer_pointer)] = i
                    worksheet_pl['A'+str(writer_pointer)].font=Font(name="Arail", size=8)
    
    
    writer_pointer=writer_pointer+2
    worksheet_pl['A'+str(writer_pointer)] = "Seller's name & address:"
    worksheet_pl['A'+str(writer_pointer)].font=font_important_content
    worksheet_pl['A'+str(writer_pointer+1)] = 'The Radio Flyer Company'+'\n'+'6515 W Grand Ave., Chicago IL 60707, USA'
    worksheet_pl['A'+str(writer_pointer+1)].font=Font(name="Arail", size=9)
    worksheet_pl['A'+str(writer_pointer+1)].alignment=Alignment(wrapText=True)
    worksheet_pl.merge_cells("A"+str(writer_pointer+1)+':B'+str(writer_pointer+3))
    
    # ---格式调整
    
    # 调整列宽
    worksheet_pl.column_dimensions['A'].width=15
    worksheet_pl.column_dimensions['B'].width=14.8
    worksheet_pl.column_dimensions['C'].width=18
    worksheet_pl.column_dimensions['D'].width=15
    worksheet_pl.column_dimensions['E'].width=1.9
    worksheet_pl.column_dimensions['F'].width=29.8
    worksheet_pl.column_dimensions['G'].width=16.8
    worksheet_pl.column_dimensions['H'].width=16.8
    worksheet_pl.column_dimensions['I'].width=16.8
    
    
    # 调整页面
    worksheet_pl.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet_pl.page_setup.fitToWidth = 1
    worksheet_pl.page_setup.fitToHeight = 0

    # 设置页眉
    worksheet_pl.oddHeader.right.text = "ORIGINAL\nPACKING SLIP"  # 文本
    worksheet_pl.oddHeader.right.size = 10  # 字号
    worksheet_pl.oddHeader.right.font = "Arial,Bold"  # 字体
    
    # 设置页边距(单位:英寸)
    worksheet_pl.page_margins = PageMargins(top=0.5, bottom=0.5, left=0.75,right=0.75) 

    return worksheet_pl


def write_excel(table_head_dict, model_dict):
    for index in range(len(table_head_dict)):
        file_name = list(table_head_dict.items())[index][0]
        temp_dict = list(table_head_dict.items())[index][1]
        temp_content_dict = list(model_dict.items())[index][1]

        # 创建一个 ExcelWriter 对象，指定 engine='openpyxl' 参数
        writer = pd.ExcelWriter(folder_path + '\\documents\\' + str(file_name) + '.xlsx', engine='openpyxl')

        # 将 DataFrame 写入 Excel 文件
        pd.DataFrame().to_excel(writer, sheet_name='Inv', index=False)
        pd.DataFrame().to_excel(writer, sheet_name='PL', index=False)
        workbook = writer.book

        worksheet_inv = writer.sheets['Inv']
        worksheet_inv = write_inv_template(worksheet_inv, file_name, temp_dict, temp_content_dict)
        worksheet_pl = writer.sheets['PL']
        worksheet_pl = write_pl_template(worksheet_pl, file_name, temp_dict, temp_content_dict)
        writer.close()


if __name__ == '__main__':
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    messagebox.showinfo("提示", "程序启动中！")

    folder_path = os.getcwd()
    # folder_path+='\\target_CCD_template'
    table_data = data_extraction(file_dir=folder_path, data_file_name='\\original data.xlsx')
    table_head_dict, model_dict = data_integration(data=table_data)
    write_excel(table_head_dict, model_dict)
    messagebox.showinfo("提示", "Completed.")
