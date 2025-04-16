import pandas as pd
from tkinter import messagebox
import tkinter as tk
import os
import openpyxl
import datetime
from pathlib import Path
from openpyxl.styles import Font,Alignment,Side, Border,PatternFill


def turn_number(num):
    if num>=0:
        return 0
    else:
        return -num

def get_data_from_po_cv(folder_dir,data_file_name):
    SKU_LIST = pd.read_excel(folder_dir+data_file_name,sheet_name='MP standard item List',header=5,usecols='A:AU',dtype=str)
    SKU_LIST = SKU_LIST[SKU_LIST['FG /Component']=='FG'].reset_index(drop=True)
    SKU_LIST.fillna('',inplace=True)

    OEM_name = pd.read_excel(folder_dir+data_file_name,sheet_name='Vendor Name',dtype=str)
    OEM_name.fillna('',inplace=True)
    
    Purchase_order_extention = pd.read_excel(folder_dir+data_file_name,sheet_name='Purchase order extension',dtype=str)
    Purchase_order_extention.fillna('',inplace=True)
    Purchase_order_extention=Purchase_order_extention[Purchase_order_extention['2nd Item Number'].isin(SKU_LIST['Model'].tolist())]
    Purchase_order_extention['Order Quantity']=Purchase_order_extention['Order Quantity'].astype(int)
    
    po_pv=Purchase_order_extention.groupby(['2nd Item Number','Vendor Number'])['Order Quantity'].sum().reset_index(name='Sum of Order Quantity')
    po_pv=pd.merge(po_pv,OEM_name[['Vendor Code','Short name (other name)']],left_on='Vendor Number',right_on='Vendor Code',how='left')
    del po_pv['Vendor Code']
    po_pv=po_pv[['Vendor Number', 
           'Short name (other name)','2nd Item Number','Sum of Order Quantity']]

    forecast = pd.read_excel(folder_dir+data_file_name,sheet_name='forecast',dtype=str,header=0,usecols='A:BB')
    forecast.fillna('',inplace=True)
    
    index=[x for x in range(12)]
    month=list(filter(lambda x:x.find('Unnamed')==-1,forecast.columns.tolist()))[1:]
    month_dist=dict(zip(index,month))
    
    forecast.columns=forecast.iloc[0]
    forecast.drop(0,inplace=True)
    new_index=list(forecast.columns)
    for i in range(0,len(forecast.columns[6:])):
        forecast[forecast.columns[6+i]]=forecast[forecast.columns[6+i]].astype(int)
        new_index[6+i]=new_index[6+i]+'_'+month_dist.get(int((i)/4))
    forecast=forecast.set_axis(new_index,axis='columns')
    forecast=forecast[forecast['Branch Plant'].apply(lambda x:x.replace(" ",'')).isin(['CHWY','FSHN','SZ','VNAM'])]
    forecast=forecast[forecast['Model'].isin(SKU_LIST['Model'].tolist())]
    
    fc_pv=forecast.groupby(['Supplier','Model'])[['Variance_Jan','Variance_Feb','Variance_Mar','Variance_Apr','Variance_May','Variance_Jun','Variance_Jul','Variance_Aug','Variance_Sep','Variance_Oct','Variance_Nov','Variance_Dec']].sum()\
        .reset_index()
    fc_pv=pd.merge(fc_pv,OEM_name[['Vendor Code','Short name (other name)']],left_on='Supplier',right_on='Vendor Code',how='left')
    del fc_pv['Vendor Code']
    fc_pv.fillna('',inplace=True)
    for col in ['Variance_Jan','Variance_Feb','Variance_Mar','Variance_Apr','Variance_May','Variance_Jun','Variance_Jul','Variance_Aug','Variance_Sep','Variance_Oct','Variance_Nov','Variance_Dec']:
        fc_pv[col]=fc_pv[col].apply(lambda x:turn_number(x))
        
    fc_pv=fc_pv[['Supplier','Short name (other name)','Model', 'Variance_Jan', 'Variance_Feb', 'Variance_Mar',
           'Variance_Apr', 'Variance_May', 'Variance_Jun', 'Variance_Jul',
           'Variance_Aug', 'Variance_Sep', 'Variance_Oct', 'Variance_Nov',
           'Variance_Dec']]
    
    res_po_fc=pd.merge(po_pv,fc_pv,left_on=['Vendor Number','2nd Item Number'],right_on=['Supplier','Model'],how='outer')
    res_po_fc.loc[res_po_fc['Vendor Number'].isna(),'Vendor Number']=res_po_fc.loc[res_po_fc['Vendor Number'].isna(),'Supplier']
    res_po_fc.loc[res_po_fc['Short name (other name)_x'].isna(),'Short name (other name)_x']=res_po_fc.loc[res_po_fc['Short name (other name)_x'].isna(),'Short name (other name)_y']
    res_po_fc.loc[res_po_fc['2nd Item Number'].isna(),'2nd Item Number']=res_po_fc.loc[res_po_fc['2nd Item Number'].isna(),'Model']
    
    del res_po_fc['Supplier'], res_po_fc['Short name (other name)_y'], res_po_fc['Model']
    res_po_fc.fillna(0,inplace=True)
    
    res_po_fc=pd.merge(res_po_fc,SKU_LIST[['Model','RB Series','Category','Item Status']],left_on=['2nd Item Number'],right_on=['Model'],how='left')
    res_po_fc=res_po_fc.drop_duplicates()
    res_po_fc=pd.merge(res_po_fc,SKU_LIST[['Model','Vendor','Item Status']],left_on=['2nd Item Number','Short name (other name)_x'],right_on=['Model','Vendor'],how='left')
    res_po_fc.loc[res_po_fc['Item Status_y'].isna(),'Item Status_y']=res_po_fc.loc[res_po_fc['Item Status_y'].isna(),'Item Status_x']
    
    res_po_fc=res_po_fc[['Vendor Number', 'Short name (other name)_x','Item Status_y','Category', 'RB Series', 'Model_x',
           'Variance_Jan', 'Variance_Feb', 'Variance_Mar',
           'Variance_Apr', 'Variance_May', 'Variance_Jun', 'Variance_Jul',
           'Variance_Aug', 'Variance_Sep', 'Variance_Oct', 'Variance_Nov',
           'Variance_Dec','Sum of Order Quantity']]
    res_po_fc.rename(columns={
        'Sum of Order Quantity':'PO QTY Receive Date   (CRD in 2024) ',
        'Category':'Key Category',
        'Model_x':'Item',
        'Item Status_y':'Item Status',
        'Short name (other name)_x':'Vendor'
    },inplace=True)
    for col in res_po_fc.columns:
        if col.find('Variance_')!=-1:
            res_po_fc.rename(columns={col:col.replace('Variance_','')},inplace=True)
    
    return res_po_fc,SKU_LIST


def get_new_rb(folder_dir,last_report_name,res_po_fc,SKU_LIST):

    rb_excel=pd.read_excel(folder_dir+last_report_name,sheet_name='Rolling RB',dtype=str)
    rb_report=pd.read_excel(folder_dir+last_report_name,sheet_name='new RB Report',dtype=str)

    # from history and new po/fc
    res_all=pd.merge(res_po_fc,rb_excel[['Vendor Number', 'Vendor', 'Item Status','Key Category', 'RB Series','Item','Adjust','2024 FG RB Cum.','remark','2023 STD CRD','2022 STD CRD','2021 STD CRD','2020 STD CRD','2019 STD CRD']],how='outer')

    # from rb report of last version
    rb_report.rename(columns={'new RB Qty':'final RB Qty'},inplace=True)
    modified_excel=pd.merge(rb_report,res_all,how='inner')
    
    modified_excel['2024 FG RB Cum.']=modified_excel['2024 FG RB Cum.'].astype(int)+modified_excel['final RB Qty'].astype(int)
    del modified_excel['CRD'],modified_excel['final RB Qty']
    

    # history and po/cv minus modified
    new_excel=pd.merge(rb_report,res_all,how='right')
    new_excel=new_excel[new_excel['CRD'].isna()]
    del new_excel['CRD'],new_excel['final RB Qty']
    
    # from new item with no po or/and cv from report
    new_added_report=rb_report[rb_report['CRD'].isna()]
    new_added_excel=pd.merge(new_added_report,SKU_LIST[['Model','RB Series','Category','Item Status']],left_on=['Item'],right_on=['Model'],how='left')
    
    new_added_excel=pd.merge(new_added_excel,SKU_LIST[['Model','Vendor','Item Status']],left_on=['Item','Vendor'],right_on=['Model','Vendor'],how='left')
    new_added_excel.loc[new_added_excel['Item Status_y'].isna(),'Item Status_y']=new_added_excel.loc[new_added_excel['Item Status_y'].isna(),'Item Status_x']
    
    new_added_excel.rename(columns={
        'Item Status_y':'Item Status',
        'Category':'Key Category'},inplace=True)
    new_added_excel=new_added_excel[['Vendor Number', 'Vendor','Item Status','Key Category', 'RB Series', 'Item','final RB Qty']]
    
    new_added_excel=pd.merge(new_added_excel,res_all,how='left')
    new_added_excel['2024 FG RB Cum.']=new_added_excel['final RB Qty'].astype(int)
    del new_added_excel['final RB Qty']

    # RESULT
    new_excel_res=pd.concat([modified_excel,new_excel])
    new_excel_res=pd.concat([new_excel_res,new_added_excel])

    sku_month=SKU_LIST[['Model','Vendor','Month']].copy()
    sku_month.loc[sku_month['Month']=='','Month']=sku_month.loc[sku_month['Month']=='','Month']=4
    
    new_excel_res=pd.merge(new_excel_res,sku_month,how='left',left_on=['Vendor','Item'],right_on=['Vendor','Model'])
    new_excel_res.loc[new_excel_res['Month'].isna(),'Month']=new_excel_res.loc[new_excel_res['Month'].isna(),'Month']=4
    new_excel_res['Month']=new_excel_res['Month'].astype(int)
    del new_excel_res['Model']
    
    new_excel_res['Bal.FG QTY after PO']=0
    new_excel_res['Bal.FG Qty(deduct 4MONTH FCST)']=0
    new_excel_res['new RB Qty']=0
    
    new_excel_res=new_excel_res[['Vendor Number',  'Vendor',
           'Item Status', 'Key Category', 'RB Series', 'Item', 
            'Jan', 'Feb', 'Mar', 'Apr','May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec',
           'PO QTY Receive Date   (CRD in 2024) ','2024 FG RB Cum.', 'Adjust','Bal.FG QTY after PO',
           'Bal.FG Qty(deduct 4MONTH FCST)', 'new RB Qty', 'remark','Month','2023 STD CRD','2022 STD CRD','2021 STD CRD','2020 STD CRD','2019 STD CRD']]
    new_excel_res=new_excel_res.sort_values(['Vendor','RB Series', 'Item'])
    new_excel_res[['PO QTY Receive Date   (CRD in 2024) ','2024 FG RB Cum.', 'Adjust','2023 STD CRD','2022 STD CRD','2021 STD CRD','2020 STD CRD','2019 STD CRD']]=\
        new_excel_res[['PO QTY Receive Date   (CRD in 2024) ','2024 FG RB Cum.', 'Adjust','2023 STD CRD','2022 STD CRD','2021 STD CRD','2020 STD CRD','2019 STD CRD']].fillna(0)
    new_excel_res[[
           'PO QTY Receive Date   (CRD in 2024) ','2024 FG RB Cum.', 'Adjust','Bal.FG QTY after PO',
           'Bal.FG Qty(deduct 4MONTH FCST)', 'new RB Qty','2023 STD CRD','2022 STD CRD','2021 STD CRD','2020 STD CRD','2019 STD CRD']]=new_excel_res[[
           'PO QTY Receive Date   (CRD in 2024) ','2024 FG RB Cum.', 'Adjust','Bal.FG QTY after PO',
           'Bal.FG Qty(deduct 4MONTH FCST)', 'new RB Qty','2023 STD CRD','2022 STD CRD','2021 STD CRD','2020 STD CRD','2019 STD CRD']].astype(int)
    
    return new_excel_res


def write_new_file(folder_dir,new_report_name,new_excel_res):
    # -----------------rb excel
    writer = pd.ExcelWriter(folder_dir+new_report_name, engine='openpyxl')
    workbook = writer.book 
    
    new_excel_res.drop('Month',axis=1).to_excel(writer, sheet_name='Rolling RB', index=False)
    worksheet_rb = writer.sheets['Rolling RB']
    
    for cell in worksheet_rb[1]:
        cell.font = Font(name="Calibri", size=12, bold=True)
        cell.alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
    for char in range(ord('G'),ord('S')):
        worksheet_rb[chr(char)+str(1)].fill = PatternFill(fill_type='solid', start_color="ddd9c4", end_color='ddd9c4')
    worksheet_rb['S1'].fill = PatternFill(fill_type='solid', start_color="95b3d7", end_color='95b3d7')
    worksheet_rb['T1'].fill = PatternFill(fill_type='solid', start_color="92d050", end_color='92d050')
    for char in range(ord('V'),ord('X')):
        worksheet_rb[chr(char)+str(1)].fill = PatternFill(fill_type='solid', start_color="fabf8f", end_color='fabf8f')
    worksheet_rb['X1'].fill = PatternFill(fill_type='solid', start_color="ffff00", end_color='ffff00')
    worksheet_rb['Z1'].fill = PatternFill(fill_type='solid', start_color="bfbfbf", end_color='bfbfbf')
    worksheet_rb['AA1'].fill = PatternFill(fill_type='solid', start_color="bfbfbf", end_color='bfbfbf')
    worksheet_rb['AB1'].fill = PatternFill(fill_type='solid', start_color="bfbfbf", end_color='bfbfbf')
    worksheet_rb['AC1'].fill = PatternFill(fill_type='solid', start_color="bfbfbf", end_color='bfbfbf')
    worksheet_rb['AD1'].fill = PatternFill(fill_type='solid', start_color="BFBFBF", end_color='BFBFBF')

    today_month=datetime.datetime.now().month
    for pointer in range(2,len(new_excel_res)+2):
        for cell in worksheet_rb[pointer]:
            cell.font = Font(name="Calibri", size=11)
        worksheet_rb['V'+str(pointer)]='=T'+str(pointer)+'-S'+str(pointer)+'+U'+str(pointer)
        month_len=new_excel_res.iloc[pointer-2]['Month']
        worksheet_rb['W'+str(pointer)]='=V'+str(pointer)+'-SUM('+chr(ord('F')+today_month)+str(pointer)+':'+chr(ord('F')+today_month+month_len-1)+str(pointer)+')'
        worksheet_rb['X'+str(pointer)]='=IF(W'+str(pointer)+'<0,-W'+str(pointer)+',0)'
    
    
    pd.DataFrame(columns=['Vendor Number','Vendor','Item','RB Qty','CRD']).to_excel(writer, sheet_name='new RB Report', index=False)
    worksheet_report = writer.sheets['new RB Report']
    for cell in worksheet_report[1]:
        cell.font = Font(name="Calibri", size=12, bold=True)
        cell.alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
        
    worksheet_report['D1'].fill = PatternFill(fill_type='solid', start_color="ffff00", end_color='ffff00')
    worksheet_report['E1'].fill = PatternFill(fill_type='solid', start_color="ffff00", end_color='ffff00')
    
    for pointer in range(2,len(new_excel_res)+2):
        for cell in worksheet_report[pointer]:
            cell.font = Font(name="Calibri", size=11)
    
    for char in range(ord('A'),ord('F')):
        worksheet_report[chr(char)+str(1)] = ''
    writer.close()

    #  -----------------python code-----------------------
    len_=len(new_excel_res)
    month_list=['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    print(datetime.datetime.today())
    print(datetime.datetime.today()+datetime.timedelta(days=46))

    date_temp_str=(datetime.datetime.today()+datetime.timedelta(days=46)).strftime("%d-%m")
    date_month=date_temp_str.split('-')[1]
    print(date_month)
    print(month_list)
    date_month=month_list[int(date_month)-1]
    
    print(date_month)
    date_temp_str=date_temp_str.split('-')[0]+'-'+date_month
    python_code="""=PY(
import pandas as pd
import datetime
report=xl("'Rolling RB'!A1:X{len_}", headers=True)
report=report[['Vendor Number','Vendor','Item','new RB Qty']]
report=report[report['new RB Qty']>0]
    
report['CRD']='{date_temp_str}'
report=report.reset_index(drop=True)
report
    """.format(len_=len_+1,date_temp_str=date_temp_str)
    with open(folder_dir+code_file_name, "w") as file:
        file.write(python_code)


if __name__ == '__main__':
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    messagebox.showinfo("提示", "程序启动中！")

    folder_dir = os.getcwd()
    file_list = Path(folder_dir).glob('*.xlsx')
    last_report_name=''
    for i in file_list:
        if i.name.find('Rolling RB Release')!=-1:
            last_report_name=i.name
        
    if len(last_report_name)==0:
        messagebox.showerror("Error", "无法找到最近历史版本！程序运行结束")
        exit(0)
    
    new_report_name='Rolling RB Release_v'+datetime.datetime.today().strftime("%m%d")+'.xlsx'
    data_file_name='SKU data base for rolling RB.xlsx'
    code_file_name="python_code.txt"
    folder_dir=folder_dir+"\\"    

    res_po_fc,SKU_LIST = get_data_from_po_cv(folder_dir=folder_dir, data_file_name=data_file_name)
    new_excel_res = get_new_rb(folder_dir=folder_dir, last_report_name=last_report_name,res_po_fc=res_po_fc, SKU_LIST=SKU_LIST)
    write_new_file(folder_dir=folder_dir, new_report_name=new_report_name, new_excel_res=new_excel_res)    
    messagebox.showinfo("提示", "Completed.")